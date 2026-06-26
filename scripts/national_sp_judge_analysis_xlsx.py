"""
Build the formatted National Singles/Pairs judge analysis Excel workbook.

Mirrors the layout in ``National SP Judge Analysis.xlsx``: group headers, raw
metrics, lookup-based deviation ratings, activity flags, and rollup columns.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scripts.national_judge_report_thresholds import (
    SINGLES_PAIRS_THRESHOLDS,
    ReportActivityThresholds,
)

LOOKUP_SHEET = "Lookup table"
RAW_SHEET = "raw_data"
ANALYSIS_SHEET = "analysis"

ANALYSIS_FIRST_DATA_ROW = 4
ANALYSIS_LAST_COL = 58
ANALYSIS_FREEZE_PANES = "B4"

# 1-based column indices (directory A–G, recent H–Q, sectionals R–Z, champs AA–AI, analysis AM–BF).
C_LAST_CHAMPS = 6
C_LAST_SECTIONALS = 7
C_RECENT_BLOCK_START = 8
C_RECENT_COMP = 9
C_RECENT_SEG = 10
C_RECENT_JS = 11
C_RECENT_RULE = 12
C_RECENT_ANOM = 13
C_RECENT_ELEM_DEV = 14
C_RECENT_ELEM_MARK = 15
C_RECENT_PCS_DEV = 16
C_RECENT_PCS_MARK = 17
C_SECT_COMP = 18
C_SECT_SEG = 19
C_SECT_JS = 20
C_SECT_RULE = 21
C_SECT_ANOM = 22
C_SECT_ELEM_DEV = 23
C_SECT_ELEM_MARK = 24
C_SECT_PCS_DEV = 25
C_SECT_PCS_MARK = 26
C_CHAMPS_COMP = 27
C_CHAMPS_SEG = 28
C_CHAMPS_JS = 29
C_CHAMPS_RULE = 30
C_CHAMPS_ANOM = 31
C_CHAMPS_ELEM_DEV = 32
C_CHAMPS_ELEM_MARK = 33
C_CHAMPS_PCS_DEV = 34
C_CHAMPS_PCS_MARK = 35
C_ACTIVITY_COMP = 36
C_ACTIVITY_SEG = 37
C_ACTIVITY_JS = 38
C_ACT_TOTAL = 39
C_ACT_QUAL = 40
C_ACT_JS = 41
C_ACT_SECT = 42
C_ACT_OVERALL = 43
C_QUAL_RULE = 44
C_QUAL_ANOM = 45
C_QUAL_ELEM_DEV = 46
C_QUAL_PCS_DEV = 47
C_QUAL_OVERALL = 48
C_SECT_PERF_RULE = 49
C_SECT_PERF_ANOM = 50
C_SECT_PERF_ELEM = 51
C_SECT_PERF_PCS = 52
C_SECT_PERF_OVERALL = 53
C_CHAMPS_PERF_RULE = 54
C_CHAMPS_PERF_ANOM = 55
C_CHAMPS_PERF_ELEM = 56
C_CHAMPS_PERF_PCS = 57
C_CHAMPS_PERF_OVERALL = 58


def _col(n: int) -> str:
    return get_column_letter(n)


HIDDEN_ANALYSIS_COLUMNS = (
    _col(C_ACTIVITY_COMP),
    _col(C_ACTIVITY_SEG),
    _col(C_ACTIVITY_JS),
)
HIDDEN_ANALYSIS_COLUMNS_NO_RULE_ERRORS = HIDDEN_ANALYSIS_COLUMNS + (
    _col(C_RECENT_RULE),
    _col(C_SECT_RULE),
    _col(C_CHAMPS_RULE),
    _col(C_QUAL_RULE),
    _col(C_SECT_PERF_RULE),
    _col(C_CHAMPS_PERF_RULE),
)

_THIN = Side(style="thin")
_NO_SIDE = Side()
_ANALYSIS_COLUMN_WIDTHS: dict[str, float] = {
    "A": 21.83,
    "B": 9.33,
    "C": 10.0,
    "D": 13.16,
    "E": 12.83,
    "F": 11.0,
    "G": 11.0,
    "H": 17.16,
    "I": 13.83,
    "J": 10.83,
    "K": 14.33,
    "L": 10.83,
    "M": 10.83,
    "N": 10.83,
    "O": 12.83,
    "P": 10.83,
    "Q": 12.83,
    "R": 15.0,
    "S": 10.83,
    "T": 13.83,
    "U": 10.83,
    "V": 10.83,
    "W": 10.83,
    "X": 10.83,
    "Y": 12.83,
    "Z": 10.83,
    "AA": 12.83,
    "AB": 15.0,
    "AC": 10.83,
    "AD": 13.83,
    "AE": 10.83,
    "AF": 10.83,
    "AG": 12.83,
    "AH": 10.83,
    "AI": 12.83,
}
for _n in range(C_ACT_TOTAL, ANALYSIS_LAST_COL + 1):
    _ANALYSIS_COLUMN_WIDTHS.setdefault(_col(_n), 10.83)
_SECTION_HEADER_LEFT_COLS = frozenset(
    {
        _col(C_RECENT_BLOCK_START),
        _col(C_RECENT_COMP),
        _col(C_SECT_COMP),
        _col(C_CHAMPS_COMP),
        _col(C_ACT_TOTAL),
        _col(C_QUAL_RULE),
        _col(C_SECT_PERF_RULE),
        _col(C_CHAMPS_PERF_RULE),
    }
)
_SECTION_HEADER_RIGHT_COLS = frozenset(
    {
        _col(C_RECENT_PCS_MARK),
        _col(C_SECT_PCS_MARK),
        _col(C_CHAMPS_PCS_MARK),
        _col(C_ACT_OVERALL),
        _col(C_QUAL_OVERALL),
        _col(C_SECT_PERF_OVERALL),
        _col(C_CHAMPS_PERF_OVERALL),
    }
)
_DATA_LEFT_BORDER_COLS = frozenset(
    {
        _col(C_RECENT_BLOCK_START),
        _col(C_ACT_TOTAL),
        _col(C_QUAL_RULE),
        _col(C_SECT_PERF_RULE),
        _col(C_CHAMPS_PERF_RULE),
    }
)
_DATA_RIGHT_BORDER_COLS = frozenset(
    {
        _col(C_RECENT_PCS_MARK),
        _col(C_SECT_PCS_MARK),
        _col(C_CHAMPS_PCS_MARK),
        _col(C_ACT_OVERALL),
        _col(C_QUAL_OVERALL),
        _col(C_SECT_PERF_OVERALL),
        _col(C_CHAMPS_PERF_OVERALL),
    }
)

_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_OVERALL_SORT_ORDER = {
    "Poor": 0,
    "Fair": 1,
    "Fair/Good": 2,
    "Good": 3,
    "N/A": 4,
}


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _to_number(value: Any) -> float | None:
    if _is_blank(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    num = _to_number(value)
    if num is None:
        return None
    return int(num)


def marking_score_rating(
    score: Any,
    *,
    fair_at: float = 1.0,
    poor_at: float = 1.3,
) -> str | None:
    """VLOOKUP-style rating for PCS/element marking scores."""
    num = _to_number(score)
    if num is None:
        return None
    if num < 0.1:
        return "N/A"
    if num < fair_at:
        return "Good"
    if num < poor_at:
        return "Fair"
    return "Poor"


def _activity_flags(
    total_comps: Any,
    competition_count: Any,
    segment_count: Any,
    junior_senior_segments: Any,
    *,
    international_flag: Any = None,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
) -> tuple[str, str, str]:
    is_int = str(international_flag or "").strip().upper() == "X"
    total = (
        "Low"
        if (_to_int(total_comps) or 0) <= thresholds.total_comps_in_role_low
        else ""
    )
    qualifying = ""
    if not is_int:
        comps = _to_int(competition_count)
        segments = _to_int(segment_count)
        if (comps is not None and comps < thresholds.competition_count_low) or (
            segments is not None and segments < thresholds.segment_count_low
        ):
            qualifying = "Low"
    junior_senior = ""
    if not is_int:
        js = _to_int(junior_senior_segments)
        if js is not None and js < thresholds.junior_senior_segment_count_low:
            junior_senior = "Low"
    return total, qualifying, junior_senior


def _sectionals_activity_flag(
    last_sectionals: Any,
    *,
    international_flag: Any = None,
    min_calendar_year: int,
) -> str:
    if str(international_flag or "").strip().upper() == "X":
        return ""
    yr = _to_int(last_sectionals)
    if yr is None or yr < min_calendar_year:
        return "Low"
    return ""


def activity_overall(*flags: str) -> str:
    count = sum(1 for value in flags if value)
    if count >= len(flags):
        return "Low"
    if count >= 2:
        return "Fair"
    if count == 1:
        return "Fair"
    return "Good"


def qualifying_rule_errors_rating(rule_errors: Any, *, activity_is_low: bool) -> str | None:
    if activity_is_low:
        return "N/A"
    count = _to_int(rule_errors)
    if count is None:
        return None
    if count >= 4:
        return "Poor"
    if count == 3:
        return "Fair"
    if count == 0:
        return "Very Good"
    return "Good"


def qualifying_anomaly_rating(
    anomaly_pct: Any,
    *,
    activity_is_low: bool,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
) -> str | None:
    if activity_is_low:
        return "N/A"
    rate = _to_number(anomaly_pct)
    if rate is None:
        return None
    if rate >= thresholds.anomaly_pct_poor:
        return "Poor"
    if rate >= thresholds.anomaly_pct_fair:
        return "Fair"
    return "Good"


def champs_rule_errors_rating(rule_errors: Any, *, has_champs: bool) -> str | None:
    if not has_champs:
        return "N/A"
    count = _to_int(rule_errors)
    if count is None:
        return None
    return {3: "Poor", 2: "Fair", 1: "Good", 0: "Very Good"}.get(count)


def sectionals_anomaly_rating(
    anomaly_pct: Any,
    *,
    has_sectionals: bool,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
) -> str | None:
    if not has_sectionals:
        return "N/A"
    rate = _to_number(anomaly_pct)
    if rate is None:
        return None
    if rate >= thresholds.champs_anomaly_pct_poor:
        return "Poor"
    if rate >= thresholds.champs_anomaly_pct_fair:
        return "Fair"
    return "Good"


def champs_anomaly_rating(
    anomaly_pct: Any,
    *,
    has_champs: bool,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
) -> str | None:
    if not has_champs:
        return "N/A"
    rate = _to_number(anomaly_pct)
    if rate is None:
        return None
    if rate >= thresholds.champs_anomaly_pct_poor:
        return "Poor"
    if rate >= thresholds.champs_anomaly_pct_fair:
        return "Fair"
    return "Good"


def performance_overall(
    *ratings: str | None,
    any_na_is_overall_na: bool = True,
    all_na_is_overall_na: bool = False,
) -> str | None:
    present = [value for value in ratings if value not in (None, "")]
    if not present:
        return None
    na_count = sum(1 for value in present if value == "N/A")
    if all_na_is_overall_na and na_count == len(present):
        return "N/A"
    if any_na_is_overall_na and na_count > 0:
        return "N/A"
    values = [value for value in present if value != "N/A"]
    if not values:
        return "N/A"
    poor = sum(1 for value in values if value == "Poor")
    good = sum(1 for value in values if value in {"Good", "Very Good"})
    fair = sum(1 for value in values if value == "Fair")
    if poor >= 2:
        return "Poor"
    if good == len(values):
        return "Good"
    if poor == 1:
        return "Fair"
    if fair > 0 and good > 0:
        return "Fair/Good"
    return "Fair"


def _analysis_sort_key(
    row: pd.Series,
    *,
    include_rule_errors: bool = True,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
    sectionals_activity_min_year: int = 0,
) -> tuple[Any, ...]:
    total_comps = _to_int(row.get("total_comps_in_role_2yr"))
    comp_count = _to_int(row.get("activity_competition_count"))
    seg_count = _to_int(row.get("activity_segment_count"))
    js_count = _to_int(row.get("activity_junior_senior_segment_count"))
    total_act, qual_act, js_act = _activity_flags(
        total_comps,
        comp_count,
        seg_count,
        js_count,
        international_flag="X" if row.get("international_judge") else None,
        thresholds=thresholds,
    )
    sect_act = _sectionals_activity_flag(
        row.get("last_sectionals_in_role"),
        international_flag="X" if row.get("international_judge") else None,
        min_calendar_year=sectionals_activity_min_year,
    )
    act_overall = activity_overall(total_act, qual_act, js_act, sect_act)
    act_low = act_overall == "Low"

    elem_rating = marking_score_rating(
        row.get("element_marking_score"),
        fair_at=thresholds.element_marking_score_fair,
        poor_at=thresholds.element_marking_score_poor,
    )
    pcs_rating = marking_score_rating(row.get("pcs_marking_score"))
    if act_low:
        elem_dev = pcs_dev = "N/A"
    else:
        elem_dev = elem_rating
        pcs_dev = pcs_rating

    qual_ratings = [
        qualifying_anomaly_rating(
            row.get("anomaly_rate_pct"),
            activity_is_low=act_low,
            thresholds=thresholds,
        ),
        elem_dev,
        pcs_dev,
    ]
    if include_rule_errors:
        qual_ratings.insert(
            0,
            qualifying_rule_errors_rating(
                row.get("total_rule_errors"), activity_is_low=act_low
            ),
        )
    overall = performance_overall(*qual_ratings)

    champs_count = _to_int(row.get("champs_competition_count")) or 0
    has_champs = champs_count > 0
    sectionals_count = _to_int(row.get("sectionals_competition_count")) or 0
    has_sectionals = sectionals_count > 0
    sectionals_ratings = [
        sectionals_anomaly_rating(
            row.get("sectionals_anomaly_rate_pct"),
            has_sectionals=has_sectionals,
            thresholds=thresholds,
        ),
        marking_score_rating(
            row.get("sectionals_element_marking_score"),
            fair_at=thresholds.element_marking_score_fair,
            poor_at=thresholds.element_marking_score_poor,
        )
        if has_sectionals
        else "N/A",
        marking_score_rating(row.get("sectionals_pcs_marking_score"))
        if has_sectionals
        else "N/A",
    ]
    if include_rule_errors:
        sectionals_ratings.insert(
            0,
            champs_rule_errors_rating(
                row.get("sectionals_total_rule_errors"), has_champs=has_sectionals
            ),
        )
    sectionals_overall = performance_overall(*sectionals_ratings)
    champs_ratings = [
        champs_anomaly_rating(
            row.get("champs_anomaly_rate_pct"),
            has_champs=has_champs,
            thresholds=thresholds,
        ),
        marking_score_rating(
            row.get("champs_element_marking_score"),
            fair_at=thresholds.element_marking_score_fair,
            poor_at=thresholds.element_marking_score_poor,
        )
        if has_champs
        else "N/A",
        marking_score_rating(row.get("champs_pcs_marking_score")) if has_champs else "N/A",
    ]
    if include_rule_errors:
        champs_ratings.insert(
            0,
            champs_rule_errors_rating(
                row.get("champs_total_rule_errors"), has_champs=has_champs
            ),
        )
    champs_overall = performance_overall(
        *champs_ratings,
        any_na_is_overall_na=False,
        all_na_is_overall_na=True,
    )

    name = str(row.get("directory_name") or "")
    rule_errors = _to_int(row.get("total_rule_errors")) if include_rule_errors else None
    anomaly = _to_number(row.get("anomaly_rate_pct"))
    return (
        _OVERALL_SORT_ORDER.get(overall, 99),
        -(rule_errors if rule_errors is not None else -1),
        -(anomaly if anomaly is not None else -1),
        _OVERALL_SORT_ORDER.get(champs_overall, 99),
        name.casefold(),
    )


def analysis_row_order(
    raw_df: pd.DataFrame,
    *,
    include_rule_errors: bool = True,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
    sectionals_activity_min_year: int = 0,
) -> pd.DataFrame:
    if raw_df.empty:
        return raw_df.copy()
    out = raw_df.copy()
    out["_analysis_sort"] = out.apply(
        lambda row: _analysis_sort_key(
            row,
            include_rule_errors=include_rule_errors,
            thresholds=thresholds,
            sectionals_activity_min_year=sectionals_activity_min_year,
        ),
        axis=1,
    )
    out = out.sort_values("_analysis_sort").drop(columns=["_analysis_sort"])
    return out.reset_index(drop=True)


def _write_lookup_sheet(
    wb: Workbook,
    *,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
) -> None:
    ws = wb.create_sheet(LOOKUP_SHEET)
    ws["A1"] = "Element Marking Score"
    ws["D1"] = "PCS Marking Score"
    t = thresholds
    rows = [
        (0, "N/A", 0, "N/A"),
        (0.1, "Good", 0.1, "Good"),
        (t.element_marking_score_fair, "Fair", 1, "Fair"),
        (t.element_marking_score_poor, "Poor", 1.3, "Poor"),
    ]
    for idx, row in enumerate(rows, start=2):
        ws.cell(idx, 1, row[0])
        ws.cell(idx, 2, row[1])
        ws.cell(idx, 4, row[2])
        ws.cell(idx, 5, row[3])


def _write_raw_sheet(wb: Workbook, raw_df: pd.DataFrame) -> None:
    ws = wb.create_sheet(RAW_SHEET)
    for col_idx, column in enumerate(raw_df.columns, start=1):
        ws.cell(1, col_idx, column)
    for row_idx, row in enumerate(raw_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            if _is_blank(value):
                continue
            ws.cell(row_idx, col_idx, value)


def _set_analysis_headers(
    ws,
    *,
    include_rule_errors: bool = True,
    performance_block_header: str = (
        "Qualifying Competition Performance (Past three years)"
    ),
    activity_column_label: str = "Qualifying Activity",
    performance_analysis_header: str = "Qualifying Performance Analysis",
    recent_period_header: str = "Last 3 years",
    junior_senior_segment_count_header: str = "# Junior/Senior Segments",
    junior_senior_activity_label: str = "Jr/Senior Activity",
    sectionals_block_header: str = "Sectionals (since 2018 for GOEs and 2022 for PCS)",
    sectionals_performance_header: str = "Sectionals Performance Analysis",
) -> None:
    ws[f"{_col(C_RECENT_BLOCK_START)}1"] = "Raw Data"
    ws[f"{_col(C_ACT_TOTAL)}1"] = "Analysis"
    ws[f"{_col(C_RECENT_BLOCK_START)}2"] = recent_period_header
    ws[f"{_col(C_RECENT_COMP)}2"] = performance_block_header
    ws[f"{_col(C_SECT_COMP)}2"] = sectionals_block_header
    ws[f"{_col(C_CHAMPS_COMP)}2"] = "Champs (since 2018 for GOEs and 2022 for PCS)"
    ws[f"{_col(C_ACT_TOTAL)}2"] = "Activity Analysis"
    ws[f"{_col(C_QUAL_RULE)}2"] = performance_analysis_header
    ws[f"{_col(C_SECT_PERF_RULE)}2"] = sectionals_performance_header
    ws[f"{_col(C_CHAMPS_PERF_RULE)}2"] = "Champs Performance Analysis"

    headers: list[str | None] = [
        "Name",
        "Int? (current or recent)",
        "USFS #",
        "US Champs (Senior) Availability",
        "Appointment Year",
        "Last Champs in Role",
        "Last Sectionals in Role",
        "Total Comps (2 years) in Role",
        "# Competitions",
        "# Segments",
        junior_senior_segment_count_header,
        "# Rule errors",
        "Anomaly %",
        "Element Dev Score",
        "Element Marking Score",
        "PCS Dev Score",
        "PCS Marking Score",
        "# Sectionals",
        "# Segments",
        junior_senior_segment_count_header,
        "# Rule errors",
        "Anomaly %",
        "Element Dev Score",
        "Element Marking Score",
        "PCS Dev Score",
        "PCS Marking Score",
        "# Champs",
        "# Segments",
        junior_senior_segment_count_header,
        "# Rule errors",
        "Anomaly %",
        "Element Dev Score",
        "Element Marking Score",
        "PCS Dev Score",
        "PCS Marking Score",
        None,
        None,
        None,
        "Total Activity",
        activity_column_label,
        junior_senior_activity_label,
        "Sectionals Activity",
        "Activity Overall",
        "Rule Errors",
        "Anomalies",
        "Element Deviation",
        "PCS Deviation",
        "Overall",
        "Rule Errors",
        "Anomalies",
        "Element Deviation",
        "PCS Deviation",
        "Overall",
        "Rule Errors",
        "Anomalies",
        "Element Deviation",
        "PCS Deviation",
        "Overall",
    ]
    for col_idx, header in enumerate(headers, start=1):
        if header is not None:
            ws.cell(3, col_idx, header)

    bold = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    group_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in (1, 2, 3):
        for col in range(1, ANALYSIS_LAST_COL + 1):
            cell = ws.cell(row, col)
            if cell.value:
                cell.font = bold
                cell.alignment = group_align if row < 3 else header_align

    ws.merge_cells(f"{_col(C_RECENT_BLOCK_START)}1:{_col(C_RECENT_PCS_MARK)}1")
    ws.merge_cells(f"{_col(C_ACT_TOTAL)}1:{_col(C_CHAMPS_PERF_OVERALL)}1")
    ws.merge_cells(f"{_col(C_RECENT_COMP)}2:{_col(C_RECENT_PCS_MARK)}2")
    ws.merge_cells(f"{_col(C_SECT_COMP)}2:{_col(C_SECT_PCS_MARK)}2")
    ws.merge_cells(f"{_col(C_CHAMPS_COMP)}2:{_col(C_CHAMPS_PCS_MARK)}2")
    ws.merge_cells(f"{_col(C_ACT_TOTAL)}2:{_col(C_ACT_OVERALL)}2")
    ws.merge_cells(f"{_col(C_QUAL_RULE)}2:{_col(C_QUAL_OVERALL)}2")
    ws.merge_cells(f"{_col(C_SECT_PERF_RULE)}2:{_col(C_SECT_PERF_OVERALL)}2")
    ws.merge_cells(f"{_col(C_CHAMPS_PERF_RULE)}2:{_col(C_CHAMPS_PERF_OVERALL)}2")

    for letter in (
        HIDDEN_ANALYSIS_COLUMNS
        if include_rule_errors
        else HIDDEN_ANALYSIS_COLUMNS_NO_RULE_ERRORS
    ):
        ws.column_dimensions[letter].hidden = True

    ws.row_dimensions[1].height = 19
    ws.row_dimensions[2].height = 19
    ws.row_dimensions[3].height = 51
    ws.sheet_format.defaultRowHeight = 16


def _border(
    *,
    left: Side | None = None,
    top: Side | None = None,
    right: Side | None = None,
    bottom: Side | None = None,
) -> Border:
    return Border(
        left=left or _NO_SIDE,
        top=top or _NO_SIDE,
        right=right or _NO_SIDE,
        bottom=bottom or _NO_SIDE,
    )


def _apply_group_header_borders(ws) -> None:
    ws[f"{_col(C_RECENT_BLOCK_START)}1"].border = _border(
        left=_THIN, top=_THIN, right=_THIN, bottom=_THIN
    )
    ws[f"{_col(C_ACT_TOTAL)}1"].border = _border(
        left=_THIN, top=_THIN, right=_THIN, bottom=_THIN
    )
    for addr in (
        _col(C_RECENT_COMP),
        _col(C_SECT_COMP),
        _col(C_CHAMPS_COMP),
        _col(C_ACT_TOTAL),
        _col(C_QUAL_RULE),
        _col(C_SECT_PERF_RULE),
    ):
        ws[f"{addr}2"].border = _border(left=_THIN, right=_THIN, bottom=_THIN)
    ws[f"{_col(C_CHAMPS_PERF_RULE)}2"].border = _border(left=_THIN, right=_THIN)


def _apply_header_row_borders(ws) -> None:
    for col in range(1, 7):
        cell = ws.cell(3, col)
        if cell.value:
            cell.border = _border(top=_THIN, bottom=_THIN)

    for col in range(7, ANALYSIS_LAST_COL + 1):
        cell = ws.cell(3, col)
        if not cell.value:
            continue
        letter = get_column_letter(col)
        cell.border = _border(
            left=_THIN if letter in _SECTION_HEADER_LEFT_COLS else None,
            top=_THIN,
            right=_THIN if letter in _SECTION_HEADER_RIGHT_COLS else None,
            bottom=_THIN,
        )


def _apply_data_row_borders(ws, *, last_row: int) -> None:
    if last_row < ANALYSIS_FIRST_DATA_ROW:
        return
    for row in range(ANALYSIS_FIRST_DATA_ROW, last_row + 1):
        for col in range(1, ANALYSIS_LAST_COL + 1):
            letter = get_column_letter(col)
            left = _THIN if letter in _DATA_LEFT_BORDER_COLS else None
            right = _THIN if letter in _DATA_RIGHT_BORDER_COLS else None
            if left or right:
                ws.cell(row, col).border = _border(left=left, right=right)


def _apply_analysis_column_widths(ws) -> None:
    for col in range(1, ANALYSIS_LAST_COL + 1):
        letter = get_column_letter(col)
        width = _ANALYSIS_COLUMN_WIDTHS.get(letter, 10.83)
        ws.column_dimensions[letter].width = width


def _apply_analysis_sheet_layout(ws, *, last_row: int) -> None:
    _apply_analysis_column_widths(ws)
    _apply_group_header_borders(ws)
    _apply_header_row_borders(ws)
    _apply_data_row_borders(ws, last_row=last_row)
    ws.freeze_panes = ANALYSIS_FREEZE_PANES


def _write_analysis_value(ws, row: int, col: int, value: Any) -> None:
    if _is_blank(value):
        return
    ws.cell(row, col, value)


def _write_analysis_formulas(
    ws,
    row: int,
    *,
    include_rule_errors: bool = True,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
    sectionals_activity_min_year: int = 0,
) -> None:
    r = row
    t = thresholds
    c = _col
    ws[f"{c(C_RECENT_ELEM_DEV)}{r}"] = (
        f"=VLOOKUP({c(C_RECENT_ELEM_MARK)}{r},'{LOOKUP_SHEET}'!A$2:B$5,2,TRUE)"
    )
    ws[f"{c(C_RECENT_PCS_DEV)}{r}"] = (
        f"=VLOOKUP({c(C_RECENT_PCS_MARK)}{r},'{LOOKUP_SHEET}'!D$2:E$5,2,TRUE)"
    )
    ws[f"{c(C_SECT_ELEM_DEV)}{r}"] = (
        f"=VLOOKUP({c(C_SECT_ELEM_MARK)}{r},'{LOOKUP_SHEET}'!A$2:B$5,2,TRUE)"
    )
    ws[f"{c(C_SECT_PCS_DEV)}{r}"] = (
        f"=VLOOKUP({c(C_SECT_PCS_MARK)}{r},'{LOOKUP_SHEET}'!D$2:E$5,2,TRUE)"
    )
    ws[f"{c(C_CHAMPS_ELEM_DEV)}{r}"] = (
        f"=VLOOKUP({c(C_CHAMPS_ELEM_MARK)}{r},'{LOOKUP_SHEET}'!A$2:B$5,2,TRUE)"
    )
    ws[f"{c(C_CHAMPS_PCS_DEV)}{r}"] = (
        f"=VLOOKUP({c(C_CHAMPS_PCS_MARK)}{r},'{LOOKUP_SHEET}'!D$2:E$5,2,TRUE)"
    )

    act_lo = c(C_ACT_TOTAL)
    act_hi = c(C_ACT_OVERALL)
    act_flags = f"{act_lo}{r}:{c(C_ACT_SECT)}{r}"
    min_year = sectionals_activity_min_year
    ws[f"{act_lo}{r}"] = (
        f'=IF({c(C_RECENT_BLOCK_START)}{r}<={t.total_comps_in_role_low},"Low","")'
    )
    ws[f"{c(C_ACT_QUAL)}{r}"] = (
        f'=IF(AND(OR({c(C_ACTIVITY_COMP)}{r}<{t.competition_count_low},'
        f'{c(C_ACTIVITY_SEG)}{r}<{t.segment_count_low}),NOT(B{r}="X")),"Low","")'
    )
    ws[f"{c(C_ACT_JS)}{r}"] = (
        f'=IF(AND(OR({c(C_ACTIVITY_JS)}{r}<{t.junior_senior_segment_count_low}),'
        f'NOT(B{r}="X")),"Low","")'
    )
    ws[f"{c(C_ACT_SECT)}{r}"] = (
        f'=IF(AND(NOT(B{r}="X"),OR({c(C_LAST_SECTIONALS)}{r}="",'
        f'{c(C_LAST_SECTIONALS)}{r}<{min_year})),"Low","")'
    )
    ws[f"{act_hi}{r}"] = (
        f'=IF(COUNTIF({act_flags},"?*")=4,"Low",'
        f'IF(COUNTIF({act_flags},"?*")>=2,"Fair",'
        f'IF(COUNTIF({act_flags},"?*")=1,"Fair","Good")))'
    )

    qual_rule = c(C_QUAL_RULE)
    qual_anom = c(C_QUAL_ANOM)
    qual_elem = c(C_QUAL_ELEM_DEV)
    qual_pcs = c(C_QUAL_PCS_DEV)
    qual_overall = c(C_QUAL_OVERALL)
    sect_rule = c(C_SECT_PERF_RULE)
    sect_anom = c(C_SECT_PERF_ANOM)
    sect_elem = c(C_SECT_PERF_ELEM)
    sect_pcs = c(C_SECT_PERF_PCS)
    sect_overall = c(C_SECT_PERF_OVERALL)
    ch_rule = c(C_CHAMPS_PERF_RULE)
    ch_anom = c(C_CHAMPS_PERF_ANOM)
    ch_elem = c(C_CHAMPS_PERF_ELEM)
    ch_pcs = c(C_CHAMPS_PERF_PCS)
    ch_overall = c(C_CHAMPS_PERF_OVERALL)

    if include_rule_errors:
        ws[f"{qual_rule}{r}"] = (
            f'=IF({act_hi}{r}="Low","N/A",'
            f'IF({c(C_RECENT_RULE)}{r}>=4,"Poor",IF({c(C_RECENT_RULE)}{r}=3,"Fair",'
            f'IF({c(C_RECENT_RULE)}{r}=0,"Very Good","Good"))))'
        )
        ws[f"{qual_overall}{r}"] = (
            f'=IF(COUNTIF({qual_rule}{r}:{qual_pcs}{r},"N/A")>0,"N/A",'
            f'IF(COUNTIF({qual_rule}{r}:{qual_pcs}{r},"Poor")>=2,"Poor",'
            f'IF(COUNTIF({qual_rule}{r}:{qual_pcs}{r},"Good")+'
            f'COUNTIF({qual_rule}{r}:{qual_pcs}{r},"Very Good")=4,"Good",'
            f'IF(COUNTIF({qual_rule}{r}:{qual_pcs}{r},"Poor")=1,"Fair",'
            f'IF(AND(COUNTIF({qual_rule}{r}:{qual_pcs}{r},"Fair")>0,'
            f'COUNTIF({qual_rule}{r}:{qual_pcs}{r},"Good")+'
            f'COUNTIF({qual_rule}{r}:{qual_pcs}{r},"Very Good")>0),'
            f'"Fair/Good","Fair")))))'
        )
        ws[f"{sect_rule}{r}"] = (
            f'=IF({c(C_SECT_COMP)}{r}=0,"N/A",'
            f'IF({c(C_SECT_RULE)}{r}=3,"Poor",IF({c(C_SECT_RULE)}{r}=2,"Fair",'
            f'IF({c(C_SECT_RULE)}{r}=1,"Good","Very Good"))))'
        )
        ws[f"{sect_overall}{r}"] = (
            f'=IF(COUNTIF({sect_rule}{r}:{sect_pcs}{r},"N/A")=4,"N/A",'
            f'IF(COUNTIF({sect_rule}{r}:{sect_pcs}{r},"Poor")>=2,"Poor",'
            f'IF(COUNTIF({sect_rule}{r}:{sect_pcs}{r},"Good")+'
            f'COUNTIF({sect_rule}{r}:{sect_pcs}{r},"Very Good")=4,"Good",'
            f'IF(COUNTIF({sect_rule}{r}:{sect_pcs}{r},"Poor")=1,"Fair",'
            f'IF(AND(COUNTIF({sect_rule}{r}:{sect_pcs}{r},"Fair")>0,'
            f'COUNTIF({sect_rule}{r}:{sect_pcs}{r},"Good")+'
            f'COUNTIF({sect_rule}{r}:{sect_pcs}{r},"Very Good")>0),'
            f'"Fair/Good","Fair")))))'
        )
        ws[f"{ch_rule}{r}"] = (
            f'=IF({c(C_CHAMPS_COMP)}{r}=0,"N/A",'
            f'IF({c(C_CHAMPS_RULE)}{r}=3,"Poor",IF({c(C_CHAMPS_RULE)}{r}=2,"Fair",'
            f'IF({c(C_CHAMPS_RULE)}{r}=1,"Good","Very Good"))))'
        )
        ws[f"{ch_overall}{r}"] = (
            f'=IF(COUNTIF({ch_rule}{r}:{ch_pcs}{r},"N/A")=4,"N/A",'
            f'IF(COUNTIF({ch_rule}{r}:{ch_pcs}{r},"Poor")>=2,"Poor",'
            f'IF(COUNTIF({ch_rule}{r}:{ch_pcs}{r},"Good")+'
            f'COUNTIF({ch_rule}{r}:{ch_pcs}{r},"Very Good")=4,"Good",'
            f'IF(COUNTIF({ch_rule}{r}:{ch_pcs}{r},"Poor")=1,"Fair",'
            f'IF(AND(COUNTIF({ch_rule}{r}:{ch_pcs}{r},"Fair")>0,'
            f'COUNTIF({ch_rule}{r}:{ch_pcs}{r},"Good")+'
            f'COUNTIF({ch_rule}{r}:{ch_pcs}{r},"Very Good")>0),'
            f'"Fair/Good","Fair")))))'
        )
    else:
        ws[f"{qual_overall}{r}"] = (
            f'=IF({act_hi}{r}="Low","N/A",'
            f'IF(COUNTIF({qual_anom}{r}:{qual_pcs}{r},"Poor")>=2,"Poor",'
            f'IF(COUNTIF({qual_anom}{r}:{qual_pcs}{r},"Good")+'
            f'COUNTIF({qual_anom}{r}:{qual_pcs}{r},"Very Good")=3,"Good",'
            f'IF(COUNTIF({qual_anom}{r}:{qual_pcs}{r},"Poor")=1,"Fair",'
            f'IF(AND(COUNTIF({qual_anom}{r}:{qual_pcs}{r},"Fair")>0,'
            f'COUNTIF({qual_anom}{r}:{qual_pcs}{r},"Good")+'
            f'COUNTIF({qual_anom}{r}:{qual_pcs}{r},"Very Good")>0),'
            f'"Fair/Good","Fair")))))'
        )
        ws[f"{sect_overall}{r}"] = (
            f'=IF(COUNTIF({sect_anom}{r}:{sect_pcs}{r},"N/A")=3,"N/A",'
            f'IF(COUNTIF({sect_anom}{r}:{sect_pcs}{r},"Poor")>=2,"Poor",'
            f'IF(COUNTIF({sect_anom}{r}:{sect_pcs}{r},"Good")+'
            f'COUNTIF({sect_anom}{r}:{sect_pcs}{r},"Very Good")=3,"Good",'
            f'IF(COUNTIF({sect_anom}{r}:{sect_pcs}{r},"Poor")=1,"Fair",'
            f'IF(AND(COUNTIF({sect_anom}{r}:{sect_pcs}{r},"Fair")>0,'
            f'COUNTIF({sect_anom}{r}:{sect_pcs}{r},"Good")+'
            f'COUNTIF({sect_anom}{r}:{sect_pcs}{r},"Very Good")>0),'
            f'"Fair/Good","Fair")))))'
        )
        ws[f"{ch_overall}{r}"] = (
            f'=IF(COUNTIF({ch_anom}{r}:{ch_pcs}{r},"N/A")=3,"N/A",'
            f'IF(COUNTIF({ch_anom}{r}:{ch_pcs}{r},"Poor")>=2,"Poor",'
            f'IF(COUNTIF({ch_anom}{r}:{ch_pcs}{r},"Good")+'
            f'COUNTIF({ch_anom}{r}:{ch_pcs}{r},"Very Good")=3,"Good",'
            f'IF(COUNTIF({ch_anom}{r}:{ch_pcs}{r},"Poor")=1,"Fair",'
            f'IF(AND(COUNTIF({ch_anom}{r}:{ch_pcs}{r},"Fair")>0,'
            f'COUNTIF({ch_anom}{r}:{ch_pcs}{r},"Good")+'
            f'COUNTIF({ch_anom}{r}:{ch_pcs}{r},"Very Good")>0),'
            f'"Fair/Good","Fair")))))'
        )

    ws[f"{qual_anom}{r}"] = (
        f'=IF({act_hi}{r}="Low","N/A",'
        f'IF({c(C_RECENT_ANOM)}{r}>={t.anomaly_pct_poor},"Poor",'
        f'IF({c(C_RECENT_ANOM)}{r}>={t.anomaly_pct_fair},"Fair","Good")))'
    )
    ws[f"{qual_elem}{r}"] = f'=IF({act_hi}{r}="Low","N/A",{c(C_RECENT_ELEM_DEV)}{r})'
    ws[f"{qual_pcs}{r}"] = f'=IF({act_hi}{r}="Low","N/A",{c(C_RECENT_PCS_DEV)}{r})'
    ws[f"{sect_anom}{r}"] = (
        f'=IF({c(C_SECT_COMP)}{r}=0,"N/A",'
        f'IF({c(C_SECT_ANOM)}{r}>={t.champs_anomaly_pct_poor},"Poor",'
        f'IF({c(C_SECT_ANOM)}{r}>={t.champs_anomaly_pct_fair},"Fair","Good")))'
    )
    ws[f"{sect_elem}{r}"] = f'=IF({c(C_SECT_COMP)}{r}=0,"N/A",{c(C_SECT_ELEM_DEV)}{r})'
    ws[f"{sect_pcs}{r}"] = f'=IF({c(C_SECT_COMP)}{r}=0,"N/A",{c(C_SECT_PCS_DEV)}{r})'
    ws[f"{ch_anom}{r}"] = (
        f'=IF({c(C_CHAMPS_COMP)}{r}=0,"N/A",'
        f'IF({c(C_CHAMPS_ANOM)}{r}>={t.champs_anomaly_pct_poor},"Poor",'
        f'IF({c(C_CHAMPS_ANOM)}{r}>={t.champs_anomaly_pct_fair},"Fair","Good")))'
    )
    ws[f"{ch_elem}{r}"] = f'=IF({c(C_CHAMPS_COMP)}{r}=0,"N/A",{c(C_CHAMPS_ELEM_DEV)}{r})'
    ws[f"{ch_pcs}{r}"] = f'=IF({c(C_CHAMPS_COMP)}{r}=0,"N/A",{c(C_CHAMPS_PCS_DEV)}{r})'


def _write_analysis_row(
    ws,
    row: int,
    record: pd.Series,
    *,
    include_rule_errors: bool = True,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
    sectionals_activity_min_year: int = 0,
) -> None:
    _write_analysis_value(ws, row, 1, record.get("directory_name"))
    if record.get("international_judge"):
        _write_analysis_value(ws, row, 2, "X")
    _write_analysis_value(ws, row, 3, record.get("mbr_number"))
    _write_analysis_value(ws, row, 4, record.get("us_champs_senior_availability"))
    _write_analysis_value(ws, row, 5, record.get("appointment_year"))
    _write_analysis_value(ws, row, C_LAST_CHAMPS, record.get("last_champs_in_role"))
    _write_analysis_value(ws, row, C_LAST_SECTIONALS, record.get("last_sectionals_in_role"))
    _write_analysis_value(ws, row, C_RECENT_BLOCK_START, record.get("total_comps_in_role_2yr"))

    _write_analysis_value(ws, row, C_RECENT_COMP, record.get("competition_count"))
    _write_analysis_value(ws, row, C_RECENT_SEG, record.get("segment_count"))
    _write_analysis_value(ws, row, C_RECENT_JS, record.get("junior_senior_segment_count"))
    if include_rule_errors:
        _write_analysis_value(ws, row, C_RECENT_RULE, record.get("total_rule_errors"))
    _write_analysis_value(ws, row, C_RECENT_ANOM, record.get("anomaly_rate_pct"))
    _write_analysis_value(ws, row, C_RECENT_ELEM_MARK, record.get("element_marking_score"))
    _write_analysis_value(ws, row, C_RECENT_PCS_MARK, record.get("pcs_marking_score"))

    sectionals_comps = _to_int(record.get("sectionals_competition_count")) or 0
    _write_analysis_value(ws, row, C_SECT_COMP, sectionals_comps)
    _write_analysis_value(ws, row, C_SECT_SEG, record.get("sectionals_segment_count"))
    _write_analysis_value(
        ws, row, C_SECT_JS, record.get("sectionals_junior_senior_segment_count")
    )
    if include_rule_errors:
        _write_analysis_value(
            ws, row, C_SECT_RULE, record.get("sectionals_total_rule_errors")
        )
    _write_analysis_value(ws, row, C_SECT_ANOM, record.get("sectionals_anomaly_rate_pct"))
    _write_analysis_value(
        ws, row, C_SECT_ELEM_MARK, record.get("sectionals_element_marking_score")
    )
    _write_analysis_value(ws, row, C_SECT_PCS_MARK, record.get("sectionals_pcs_marking_score"))

    champs_comps = _to_int(record.get("champs_competition_count")) or 0
    _write_analysis_value(ws, row, C_CHAMPS_COMP, champs_comps)
    _write_analysis_value(ws, row, C_CHAMPS_SEG, record.get("champs_segment_count"))
    _write_analysis_value(
        ws, row, C_CHAMPS_JS, record.get("champs_junior_senior_segment_count")
    )
    if include_rule_errors:
        _write_analysis_value(ws, row, C_CHAMPS_RULE, record.get("champs_total_rule_errors"))
    _write_analysis_value(ws, row, C_CHAMPS_ANOM, record.get("champs_anomaly_rate_pct"))
    _write_analysis_value(
        ws, row, C_CHAMPS_ELEM_MARK, record.get("champs_element_marking_score")
    )
    _write_analysis_value(ws, row, C_CHAMPS_PCS_MARK, record.get("champs_pcs_marking_score"))

    _write_analysis_value(ws, row, C_ACTIVITY_COMP, record.get("activity_competition_count"))
    _write_analysis_value(ws, row, C_ACTIVITY_SEG, record.get("activity_segment_count"))
    _write_analysis_value(
        ws, row, C_ACTIVITY_JS, record.get("activity_junior_senior_segment_count")
    )

    _write_analysis_formulas(
        ws,
        row,
        include_rule_errors=include_rule_errors,
        thresholds=thresholds,
        sectionals_activity_min_year=sectionals_activity_min_year,
    )


def _rating_text_rules(top_left: str) -> list[FormulaRule]:
    """Good/Very Good green, Fair yellow, Poor red (matches manual workbook)."""
    col = top_left.rstrip("0123456789")
    row = top_left[len(col) :]
    cell = f"{col}{row}"
    return [
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("Poor",{cell})))'],
            stopIfTrue=True,
            fill=_FILL_RED,
        ),
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("Fair/Good",{cell})))'],
            stopIfTrue=True,
            fill=_FILL_YELLOW,
        ),
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("Good",{cell})))'],
            stopIfTrue=True,
            fill=_FILL_GREEN,
        ),
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("Fair",{cell})))'],
            stopIfTrue=True,
            fill=_FILL_YELLOW,
        ),
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("Low",{cell})))'],
            stopIfTrue=True,
            fill=_FILL_RED,
        ),
    ]


def _anomaly_rules(
    top_left: str,
    *,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
    poor_at: float | None = None,
    fair_at: float | None = None,
) -> list:
    col = top_left.rstrip("0123456789")
    row = top_left[len(col) :]
    cell = f"{col}{row}"
    poor = poor_at if poor_at is not None else thresholds.anomaly_pct_poor
    fair = fair_at if fair_at is not None else thresholds.anomaly_pct_fair
    return [
        FormulaRule(
            formula=[f"AND(ISNUMBER({cell}),{cell}>={poor})"],
            stopIfTrue=True,
            fill=_FILL_RED,
        ),
        FormulaRule(
            formula=[f"AND(ISNUMBER({cell}),{cell}>={fair},{cell}<{poor})"],
            stopIfTrue=True,
            fill=_FILL_YELLOW,
        ),
        FormulaRule(
            formula=[f"AND(ISNUMBER({cell}),{cell}<{fair})"],
            stopIfTrue=True,
            fill=_FILL_GREEN,
        ),
    ]


def _apply_analysis_conditional_formatting(
    ws,
    *,
    last_row: int,
    include_rule_errors: bool = True,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
    sectionals_activity_min_year: int = 0,
) -> None:
    if last_row < ANALYSIS_FIRST_DATA_ROW:
        return
    first = ANALYSIS_FIRST_DATA_ROW
    cf = ws.conditional_formatting

    def addr(col: str) -> str:
        return f"{col}{first}:{col}{last_row}"

    cf.add(
        addr("D"),
        FormulaRule(
            formula=[f'LEFT(D{first},LEN("Available"))="Available"'],
            stopIfTrue=True,
            fill=_FILL_GREEN,
        ),
    )
    cf.add(
        addr("D"),
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("Unavailable",D{first})))'],
            stopIfTrue=True,
            fill=_FILL_RED,
        ),
    )
    cf.add(
        addr("D"),
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("Didn\'t reply",D{first})))'],
            stopIfTrue=True,
            fill=_FILL_YELLOW,
        ),
    )

    t = thresholds
    min_year = sectionals_activity_min_year
    sect_col = _col(C_LAST_SECTIONALS)
    cf.add(
        addr(sect_col),
        FormulaRule(
            formula=[f'OR({sect_col}{first}="",{sect_col}{first}<{min_year})'],
            stopIfTrue=True,
            fill=_FILL_RED,
        ),
    )
    h_col = _col(C_RECENT_BLOCK_START)
    cf.add(
        addr(h_col),
        FormulaRule(
            formula=[
                f"AND(ISNUMBER({h_col}{first}),{h_col}{first}<{t.competition_count_low})"
            ],
            stopIfTrue=True,
            fill=_FILL_RED,
        ),
    )
    for col_idx, threshold in (
        (C_RECENT_COMP, t.competition_count_low),
        (C_RECENT_SEG, t.segment_count_low),
        (C_RECENT_JS, t.junior_senior_segment_count_low),
    ):
        col_letter = _col(col_idx)
        cf.add(
            addr(col_letter),
            FormulaRule(
                formula=[
                    f"AND(ISNUMBER({col_letter}{first}),"
                    f"{col_letter}{first}<{threshold},NOT(B{first}=\"X\"))"
                ],
                stopIfTrue=True,
                fill=_FILL_RED,
            ),
        )

    if include_rule_errors:
        cf.add(
            addr(_col(C_RECENT_RULE)),
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(t.qualifying_rule_errors_poor)],
                fill=_FILL_RED,
            ),
        )
        cf.add(
            addr(_col(C_SECT_RULE)),
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(t.champs_rule_errors_poor)],
                fill=_FILL_RED,
            ),
        )
        cf.add(
            addr(_col(C_CHAMPS_RULE)),
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(t.champs_rule_errors_poor)],
                fill=_FILL_RED,
            ),
        )

    for col, fair_at, poor_at in (
        (C_RECENT_ANOM, t.anomaly_pct_fair, t.anomaly_pct_poor),
        (C_SECT_ANOM, t.champs_anomaly_pct_fair, t.champs_anomaly_pct_poor),
        (C_CHAMPS_ANOM, t.champs_anomaly_pct_fair, t.champs_anomaly_pct_poor),
    ):
        for rule in _anomaly_rules(
            f"{_col(col)}{first}",
            thresholds=thresholds,
            poor_at=poor_at,
            fair_at=fair_at,
        ):
            cf.add(addr(_col(col)), rule)

    for col in (
        C_RECENT_ELEM_DEV,
        C_RECENT_PCS_DEV,
        C_SECT_ELEM_DEV,
        C_SECT_PCS_DEV,
        C_CHAMPS_ELEM_DEV,
        C_CHAMPS_PCS_DEV,
    ):
        for rule in _rating_text_rules(f"{_col(col)}{first}"):
            cf.add(addr(_col(col)), rule)

    summary_cols = (
        C_ACT_OVERALL,
        C_QUAL_OVERALL,
        C_SECT_PERF_OVERALL,
        C_CHAMPS_PERF_OVERALL,
    )
    for col in summary_cols:
        for rule in _rating_text_rules(f"{_col(col)}{first}"):
            cf.add(addr(_col(col)), rule)


def write_national_sp_judge_analysis_xlsx(
    raw_df: pd.DataFrame,
    output_path: Path,
    *,
    include_rule_errors: bool = True,
    thresholds: ReportActivityThresholds = SINGLES_PAIRS_THRESHOLDS,
    performance_block_header: str = (
        "Qualifying Competition Performance (Past three years)"
    ),
    activity_column_label: str = "Qualifying Activity",
    performance_analysis_header: str = "Qualifying Performance Analysis",
    recent_period_header: str = "Last 3 years",
    junior_senior_segment_count_header: str = "# Junior/Senior Segments",
    junior_senior_activity_label: str = "Jr/Senior Activity",
    sectionals_block_header: str = "Sectionals (since 2018 for GOEs and 2022 for PCS)",
    sectionals_performance_header: str = "Sectionals Performance Analysis",
    sectionals_activity_min_year: int = 0,
) -> None:
    """Write formatted workbook with analysis, raw data, and lookup sheets."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(ANALYSIS_SHEET)
    _set_analysis_headers(
        ws,
        include_rule_errors=include_rule_errors,
        performance_block_header=performance_block_header,
        activity_column_label=activity_column_label,
        performance_analysis_header=performance_analysis_header,
        recent_period_header=recent_period_header,
        junior_senior_segment_count_header=junior_senior_segment_count_header,
        junior_senior_activity_label=junior_senior_activity_label,
        sectionals_block_header=sectionals_block_header,
        sectionals_performance_header=sectionals_performance_header,
    )

    _write_raw_sheet(wb, raw_df)
    _write_lookup_sheet(wb, thresholds=thresholds)

    ordered = analysis_row_order(
        raw_df,
        include_rule_errors=include_rule_errors,
        thresholds=thresholds,
        sectionals_activity_min_year=sectionals_activity_min_year,
    )
    for offset, (_, record) in enumerate(ordered.iterrows()):
        row = ANALYSIS_FIRST_DATA_ROW + offset
        _write_analysis_row(
            ws,
            row,
            record,
            include_rule_errors=include_rule_errors,
            thresholds=thresholds,
            sectionals_activity_min_year=sectionals_activity_min_year,
        )

    last_row = ANALYSIS_FIRST_DATA_ROW + len(ordered) - 1 if len(ordered) else ANALYSIS_FIRST_DATA_ROW
    _apply_analysis_sheet_layout(ws, last_row=last_row)
    _apply_analysis_conditional_formatting(
        ws,
        last_row=last_row,
        include_rule_errors=include_rule_errors,
        thresholds=thresholds,
        sectionals_activity_min_year=sectionals_activity_min_year,
    )
    wb.save(output_path)
