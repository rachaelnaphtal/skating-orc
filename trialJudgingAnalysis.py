import openpyxl.worksheet
from pypdf import PdfReader
import pandas as pd
import re
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule

import judgingParsing
from judgingParsing import parse_scores
from judgingParsing import printToExcel
import downloadResults
from downloadResults import make_competition_summary_page
from gcp_interactions_helper import read_file_from_gcp
from io import BytesIO
import gcp_interactions_helper
import streamlit as st
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
    Color,
)

state_options = [
    "In Range",
    "Low- out of range with allowance",
    "Low- out of range",
    "Low- thrown out",
    "High- out of range with allowance",
    "High- out of range",
    "High- thrown out",
]


def get_judges_name_from_sheet(tj_sheet_path, use_gcp=False):
    if use_gcp:
        workbook = openpyxl.load_workbook(
            read_file_from_gcp(tj_sheet_path), data_only=True
        )
    else:
        workbook = openpyxl.load_workbook(tj_sheet_path, data_only=True)
    judges = []
    for sheet in workbook:
        if not sheet.title == "1":
            continue
        found_end = False
        judge_number = 0
        while not found_end:
            judge_name = sheet.cell(row=4 + judge_number, column=2).value
            if judge_name == "" or judge_name is None:
                return judges
            judges.append(judge_name)
            judge_number += 1


def process_trial_judge_sheet(pdf_path, num_judges=7, use_gcp=False, judge_names=[]):
    if use_gcp:
        workbook = openpyxl.load_workbook(read_file_from_gcp(pdf_path), data_only=True)
    else:
        workbook = openpyxl.load_workbook(pdf_path, data_only=True)
    tj_scores = {}
    pcs_scores = {}
    for sheet in workbook:
        if not re.match(r"^(\d+)", sheet.title):
            continue
        skater_name = sheet.cell(
            row=2,
            column=2,
        ).value
        if skater_name is None or skater_name == 0:
            continue
        skater_name = skater_name.replace("ﬁ", "fi")
        tj_scores[skater_name] = {}
        pcs_scores[skater_name] = {}
        for judge_number in range(num_judges):
            judge_name = sheet.cell(row=4 + judge_number, column=2).value
            if judge_name not in judge_names:
                err = f"Judge '{judge_name}' found in sheet but is not listed in names {judge_names}"
                st.error(err)
                raise Exception(err)
            judge_scores = []
            for i in range(13):
                score = sheet.cell(row=4 + judge_number, column=2 + i).value
                if score is None:
                    break
                judge_scores.append(score)
            tj_scores[skater_name][judge_name] = judge_scores
            pcs_scores_list = []
            for i in range(3):
                score = sheet.cell(row=4 + judge_number, column=16 + i).value
                pcs_scores_list.append(score)
            pcs_scores[skater_name][judge_name] = pcs_scores_list
    return tj_scores, pcs_scores


def analyze_trial_judges(
    tj_xlsx_path, workbook, pdf_path, judges, pdf_number, tj_filter, use_gcp=False, per_trial_judge_workbook_dict={}
):
    if len(tj_filter) == 0:
        tj_filter = judges
    (elements_per_skater, pcs_per_skater, _, event_name) = parse_scores(
        pdf_path, use_gcp=use_gcp
    )
    tj_scores, pcs_scores = process_trial_judge_sheet(
        tj_xlsx_path, num_judges=len(judges), use_gcp=use_gcp, judge_names=judges
    )
    element_errors = []
    total_element_count = 0
    deviation_totals = {judge: 0 for judge in judges}
    all_elements = []
    for skater in tj_scores:
        if skater not in elements_per_skater:
            err = f"Skater {skater} in {event_name} was found in xlsx but not pdf. Was the name spelled correctly on the uploaded sheet? Found skaters on pdf: {elements_per_skater.keys()}"
            st.error(err)
            raise Exception(err)
        for element_details in elements_per_skater[skater]:
            allScores = element_details["Scores"]
            avg = sum(allScores) / len(allScores)
            total_element_count += 1

            element_name_no_level = element_details["Element"]
            if element_name_no_level[-1].isdigit():
                element_name_no_level = element_name_no_level[:-1]

            judge_number = 1
            for judge in tj_scores[skater]:
                if judge not in tj_filter:
                    judge_number += 1
                    continue
                if len(tj_scores[skater][judge]) - 1 < element_details["Number"]:
                    err = f"Score length error on {skater} and judge {judge} in event {event_name}. Are they missing a GOE?"
                    st.error(err)
                    raise Exception(err)
                goe = tj_scores[skater][judge][element_details["Number"]]
                if not isinstance(goe, int):
                    err = f"GOE of {goe} is not an integer. Skater: {skater}, Judge: {judge}, Event: {event_name}, Element: {element_details['Number']}"
                    st.error(err)
                    raise Exception(err)
                deviation = goe - avg
                deviation_totals[judge] += abs(deviation)
                if abs(deviation) >= 2:
                    element_errors.append(
                        {
                            "Skater": skater,
                            "Element": element_details["Element"],
                            "Judge Name": judge,
                            "Judge Number": judge_number,
                            "Judge Score": goe,
                            "Panel Average": avg,
                            "Deviation": deviation,
                            "Type": "Deviation",
                        }
                    )
                # Track how often thrown out
                state = get_relevant_out_of_range_state(
                    goe, min(allScores), max(allScores)
                )
                all_elements.append(
                    {
                        "Skater": skater,
                        "Event": event_name,
                        "Element": element_name_no_level,
                        "Element Type": categorizeElement(element_name_no_level),
                        "Panel Average": avg,
                        "Judge Name": judge,
                        "Judge Number": judge_number,
                        "Type": state,
                    }
                )

                judge_number += 1

    pcs_errors, all_pcs = add_pcs_errors(
        pcs_per_skater, pcs_scores, tj_filter, event_name
    )
    printToExcel(
        workbook, event_name, judges, [], element_errors, pcs_errors, pdf_number
    )
    
    #Add sheets to each trial judge sheet if requested
    for trial_judge in per_trial_judge_workbook_dict:
        filtered_element_errors = []
        for error in element_errors:
            if error["Judge Name"] == trial_judge:
                filtered_element_errors.append(error)
        
        filtered_pcs_errors = []
        for error in pcs_errors:
            if error["Judge Name"] == trial_judge:
                filtered_pcs_errors.append(error)

        printToExcel(
        per_trial_judge_workbook_dict[trial_judge], event_name, judges, [], filtered_element_errors, filtered_pcs_errors, pdf_number
        )

    total_errors = judgingParsing.count_total_errors_per_judge(
        judges, [], element_errors, pcs_errors
    )
    num_starts = len(elements_per_skater)
    allowed_errors = judgingParsing.get_allowed_errors(num_starts)

    # Other analysis outside ORC
    average_deviations = {
        judge: deviation_totals[judge] / total_element_count
        for judge in deviation_totals
    }
    all_elements_df = pd.DataFrame.from_dict(all_elements)
    all_pcs_df = pd.DataFrame.from_dict(all_pcs)

    return (
        event_name,
        total_errors,
        num_starts,
        allowed_errors,
        average_deviations,
        all_elements_df,
        all_pcs_df,
    )


def get_relevant_out_of_range_state(tj_score, min_panel, max_panel, allowance=1):
    state = "In Range"
    if (min_panel == max_panel) and (min_panel == tj_score):
        state = "In Range"
    elif tj_score < min_panel - allowance:
        state = "Low- out of range with allowance"
    elif tj_score < min_panel:
        state = "Low- out of range"
    elif tj_score == min_panel:
        state = "Low- thrown out"
    elif tj_score > max_panel + allowance:
        state = "High- out of range with allowance"
    elif tj_score > max_panel:
        state = "High- out of range"
    elif tj_score == max_panel:
        state = "High- thrown out"
    return state


def add_pcs_errors(pcs_per_skater, tj_pcs_scores, tj_filter, event_name):
    errors = []
    all_pcs = []
    for skater in tj_pcs_scores:
        for pcs_mark in pcs_per_skater[skater]:
            allScores = pcs_mark["Scores"]
            avg = sum(allScores) / len(allScores)
            judgeNumber = 1
            for judge in tj_pcs_scores[skater]:
                if judge not in tj_filter:
                    judgeNumber += 1
                    continue
                tj_score = tj_pcs_scores[skater][judge][
                    get_component_number(pcs_mark["Component"])
                ]
                if not (isinstance(tj_score, float) or isinstance(tj_score, int)):
                    try:
                        if "." in tj_score:
                            tj_score = float(tj_score)
                        else:
                            tj_score = int(tj_score)
                    except:
                        err = f"PCS score {tj_score} found that isn't a number. Skater: {skater}, Judge:{judge}, Event:{event_name} Component: {pcs_mark['Component']} Scores:{tj_pcs_scores[skater][judge]}"
                        st.error(err)
                        raise Exception(err)
                deviation = tj_score - avg
                all_pcs.append(
                    {
                        "Skater": skater,
                        "Event": event_name,
                        "Component": pcs_mark["Component"],
                        "Panel Average": avg,
                        "Judge Name": judge,
                        "Judge Number": judgeNumber,
                        "Score": tj_score,
                        "Type": get_relevant_out_of_range_state(
                            tj_score, min(allScores), max(allScores), allowance=0.5
                        ),
                    }
                )
                if abs(deviation) >= 1.5:
                    errors.append(
                        {
                            "Skater": skater,
                            "Judge Number": judgeNumber,
                            "Judge Name": judge,
                            "Judge Score": tj_score,
                            "Deviation": deviation,
                            "Component": pcs_mark["Component"],
                            "Type": "Deviation",
                        }
                    )
                judgeNumber += 1
    return errors, all_pcs


def get_component_number(name):
    if name == "Skating Skills":
        return 2
    if name == "Composition":
        return 0
    if name == "Presentation":
        return 1


# Return type of element
def categorizeElement(element):
    element = element.replace("<", "")
    if element[-1] == "V":
        element = element[:-1]
    if element[-1].isdigit():
        element = element[:-1]
    if element == "PB":
        return "Pivoting Block"
    if element[-1] == "B":
        element = element[:-1]

    synchro_dict = {
        "Pa": "Pair Element",
        "TrE": "Travelling Element",
        "ME": "Moves Element",
        "TwE": "Twizzle Element",
        "AL": "Artistic",
        "AC": "Artistic",
        "AW": "Artistic",
        "AB": "Artistic",
        "L": "Linear/Rotating",
        "C": "Linear/Rotating",
        "B": "Linear/Rotating",
        "W": "Linear/Rotating",
        "Cr": "Creative",
        "GL": "Group Lift",
    }

    if element in ["FiDs", "FoDS", "BiDs", "BoDs"]:
        return "Death Spiral"
    elif element.endswith("Tw"):
        return "Twist"
    elif element in ["PSp", "PCoSp"]:
        return "Pairs Spin"
    elif element in ["StSq" or "ChSq"]:
        return element
    elif element.endswith("Li"):
        return "Lift"
    elif element.endswith("Sp"):
        return "Spin"
    elif element.endswith("Th"):
        return "Lift"
    elif element[0] in ["1", "2", "3", "4"] and element[1] in ["A", "S", "T", "L", "F"]:
        return "Jump"
    elif element.endswith("+pi") or element == "I":
        return "Intersection"
    elif element.startswith("NHE"):
        return "No Hold Element"
    elif element in synchro_dict:
        return synchro_dict[element]
    return element


def make_analysis_cover_sheet(workbook):
    sheet = workbook.create_sheet("Overview")

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 150
    sheet.row_dimensions[3].height = 40
    sheet.row_dimensions[15].height = 40

    bold = Font(bold=True)

    sheet.cell(1, 1, value="Overview").font = Font(bold=True, size=18)
    sheet.merge_cells('A3:B4')
    sheet.cell(3, 1, value="This workbook contains additional analysis of the trial judging data, specifically related to deviations. \n The first six sheets show the percentage of GOEs or PCS of each type that are extremes related to the judging panel. In general, higher numbers are worse. The first six columns show the absolute numbers and the final three show the percentages of the total.")
    sheet.cell(3,1).alignment = Alignment(wrap_text=True)
    
    sheet.cell(7, 1, value="Definitions:").font = bold
    sheet.cell(8, 1, value="Within 1/ Within .5:").font = bold
    sheet.cell(9, 1, value="Out of Range:").font = bold
    sheet.cell(10, 1, value="Thrown out:").font = bold
    sheet.cell(11, 1, value="Low vs High:").font = bold

    sheet.cell(8, 2, value="The number of GOE/PCS that are outside the range of the panel +/-1 on each side. For example, if the minimum score on the panel is 1 then any scores under 0 count as two low on that sheet. For PCS it is for within +/- .5 of the panel.")
    sheet.cell(9, 2, value="The number of scores that are outside the range of the official judging panel.")
    sheet.cell(10, 2, value="The number of scores that would be the extremes of the panel or are outside the range of the panel. For example, if the high score on the panel is +2 then all trial judge scores that are +2 or higher would be counted. Note: if the full panel scores the same thing and the trial judge does as well then it will not count as thrown out.")
    sheet.cell(11, 2, value="Low refers to the trial judge being lower than the panel. The total is the sum of low and high.")
    sheet.cell(8,2).alignment = Alignment(wrap_text=True)
    sheet.cell(9,2).alignment = Alignment(wrap_text=True)
    sheet.cell(10,2).alignment = Alignment(wrap_text=True)
    sheet.cell(11,2).alignment = Alignment(wrap_text=True)

    sheet.merge_cells('A15:B15')
    sheet.cell(14, 1, value="Average GOE Deviations").font = bold
    sheet.cell(15, 1, value="This calculates the average amount that the trial judge GOE deviates from the mean score. It is calculated by summing the absolute values of the trial judge GOE vs mean score from the judging panel for that element and dividing by the number of elements.")
    sheet.cell(15,1).alignment = Alignment(wrap_text=True)

def make_extra_analysis_sheet(
    excel_path, all_element_df, all_pcs_df, average_deviations, use_gcp=False
):
    # Make additional analysis document
    extra_info_path = excel_path.replace(".xlsx", "_AdditionalAnalysis.xlsx")
    excel_buffer = BytesIO()
    if not use_gcp:
        excel_buffer = extra_info_path
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        # Add cover sheet
        make_analysis_cover_sheet(writer.book)

        # Analyze elements
        all_el_df = (
            all_element_df.groupby(["Judge Name", "Element Type", "Type"])
            .size()
            .unstack(fill_value=0)
        )
        for state in state_options:
            if not state in all_el_df:
                all_pcs_df[state] = 0
        all_el_df["Total # Elements"] = (
            all_el_df["High- out of range"]
            + all_el_df["High- thrown out"]
            + all_el_df["High- out of range with allowance"]
            + all_el_df["In Range"]
            + all_el_df["Low- out of range"]
            + all_el_df["Low- thrown out"]
            + all_el_df["Low- out of range with allowance"]
        )
        all_el_df["Low- thrown out"] = (
            all_el_df["Low- out of range with allowance"]
            + all_el_df["Low- out of range"]
            + all_el_df["Low- thrown out"]
        )
        all_el_df["High- thrown out"] = (
            all_el_df["High- out of range with allowance"]
            + all_el_df["High- out of range"]
            + all_el_df["High- thrown out"]
        )
        all_el_df["Low- out of range"] = (
            all_el_df["Low- out of range with allowance"]
            + all_el_df["Low- out of range"]
        )
        all_el_df["High- out of range"] = (
            all_el_df["High- out of range with allowance"]
            + all_el_df["High- out of range"]
        )
        all_el_df = all_el_df.reset_index()

        with_allowance_df = all_el_df[
            [
                "Judge Name",
                "Element Type",
                "Low- out of range with allowance",
                "High- out of range with allowance",
                "Total # Elements",
            ]
        ].copy()
        with_allowance_df["Low %"] = (
            with_allowance_df["Low- out of range with allowance"]
            / with_allowance_df["Total # Elements"]
        )
        with_allowance_df["High %"] = (
            with_allowance_df["High- out of range with allowance"]
            / with_allowance_df["Total # Elements"]
        )
        with_allowance_df["Out %"] = (
            with_allowance_df["High- out of range with allowance"]
            + with_allowance_df["Low- out of range with allowance"]
        ) / with_allowance_df["Total # Elements"]
        with_allowance_df.to_excel(
            writer, sheet_name="GOE- within 1", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["GOE- within 1"])

        out_of_range_df = all_el_df[
            [
                "Judge Name",
                "Element Type",
                "Low- out of range",
                "High- out of range",
                "Total # Elements",
            ]
        ].copy()
        out_of_range_df["Low %"] = (
            out_of_range_df["Low- out of range"] / out_of_range_df["Total # Elements"]
        )
        out_of_range_df["High %"] = (
            out_of_range_df["High- out of range"] / out_of_range_df["Total # Elements"]
        )
        out_of_range_df["Out %"] = (
            out_of_range_df["High- out of range"] + out_of_range_df["Low- out of range"]
        ) / out_of_range_df["Total # Elements"]
        out_of_range_df.to_excel(
            writer, sheet_name="GOE- Out of Range", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["GOE- Out of Range"])

        thrown_out_df = all_el_df[
            [
                "Judge Name",
                "Element Type",
                "Low- thrown out",
                "High- thrown out",
                "Total # Elements",
            ]
        ].copy()
        thrown_out_df["Low %"] = (
            thrown_out_df["Low- thrown out"] / thrown_out_df["Total # Elements"]
        )
        thrown_out_df["High %"] = (
            thrown_out_df["High- thrown out"] / thrown_out_df["Total # Elements"]
        )
        thrown_out_df["Out %"] = (
            thrown_out_df["High- thrown out"] + thrown_out_df["Low- thrown out"]
        ) / thrown_out_df["Total # Elements"]
        thrown_out_df.to_excel(
            writer, sheet_name="GOE-Thrown out", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["GOE-Thrown out"])

        ## PCS
        all_pcs_df = (
            all_pcs_df.groupby(["Judge Name", "Component", "Type"])
            .size()
            .unstack(fill_value=0)
        )
        for state in state_options:
            if not state in all_pcs_df:
                all_pcs_df[state] = 0

        all_pcs_df["Total # Components"] = (
            all_pcs_df["High- out of range"]
            + all_pcs_df["High- thrown out"]
            + all_pcs_df["High- out of range with allowance"]
            + all_pcs_df["In Range"]
            + all_pcs_df["Low- out of range"]
            + all_pcs_df["Low- thrown out"]
            + all_pcs_df["Low- out of range with allowance"]
        )
        all_pcs_df["Low- thrown out"] = (
            all_pcs_df["Low- out of range with allowance"]
            + all_pcs_df["Low- out of range"]
            + all_pcs_df["Low- thrown out"]
        )
        all_pcs_df["Low- thrown out"] = (
            all_pcs_df["High- out of range with allowance"]
            + all_pcs_df["High- out of range"]
            + all_pcs_df["High- thrown out"]
        )
        all_pcs_df["Low- out of range"] = (
            all_pcs_df["Low- out of range with allowance"]
            + all_pcs_df["Low- out of range"]
        )
        all_pcs_df["Low- out of range"] = (
            all_pcs_df["High- out of range with allowance"]
            + all_pcs_df["High- out of range"]
        )
        all_pcs_df = all_pcs_df.reset_index()

        pcs_with_allowance_df = all_pcs_df[
            [
                "Judge Name",
                "Component",
                "Low- out of range with allowance",
                "High- out of range with allowance",
                "Total # Components",
            ]
        ].copy()
        pcs_with_allowance_df["Low %"] = (
            pcs_with_allowance_df["Low- out of range with allowance"]
            / pcs_with_allowance_df["Total # Components"]
        )
        pcs_with_allowance_df["High %"] = (
            pcs_with_allowance_df["High- out of range with allowance"]
            / pcs_with_allowance_df["Total # Components"]
        )
        pcs_with_allowance_df["Out %"] = (
            pcs_with_allowance_df["High- out of range with allowance"]
            + pcs_with_allowance_df["Low- out of range with allowance"]
        ) / pcs_with_allowance_df["Total # Components"]
        pcs_with_allowance_df.to_excel(
            writer, sheet_name="PCS- within .5", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["PCS- within .5"])

        pcs_out_of_range_df = all_pcs_df[
            [
                "Judge Name",
                "Component",
                "Low- out of range",
                "High- out of range",
                "Total # Components",
            ]
        ].copy()
        pcs_out_of_range_df["Low %"] = (
            pcs_out_of_range_df["Low- out of range"]
            / pcs_out_of_range_df["Total # Components"]
        )
        pcs_out_of_range_df["High %"] = (
            pcs_out_of_range_df["High- out of range"]
            / pcs_out_of_range_df["Total # Components"]
        )
        pcs_out_of_range_df["Out %"] = (
            pcs_out_of_range_df["High- out of range"]
            + pcs_out_of_range_df["Low- out of range"]
        ) / pcs_out_of_range_df["Total # Components"]
        pcs_out_of_range_df.to_excel(
            writer, sheet_name="PCS- Out of Range", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["PCS- Out of Range"])

        pcs_thrown_out_df = all_pcs_df[
            [
                "Judge Name",
                "Component",
                "Low- thrown out",
                "High- thrown out",
                "Total # Components",
            ]
        ].copy()
        pcs_thrown_out_df["Low %"] = (
            pcs_thrown_out_df["Low- thrown out"]
            / pcs_thrown_out_df["Total # Components"]
        )
        pcs_thrown_out_df["High %"] = (
            pcs_thrown_out_df["High- thrown out"]
            / pcs_thrown_out_df["Total # Components"]
        )
        pcs_thrown_out_df["Out %"] = (
            pcs_thrown_out_df["High- thrown out"] + pcs_thrown_out_df["Low- thrown out"]
        ) / pcs_thrown_out_df["Total # Components"]
        pcs_thrown_out_df.to_excel(
            writer, sheet_name="PCS-Thrown out", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["PCS-Thrown out"])

        average_deviation_df = pd.DataFrame.from_records(average_deviations)
        average_deviation_df.to_excel(writer, sheet_name="Average GOE deviations")

        for sheet in writer.sheets:
            if sheet !="Overview":
                judgingParsing.autofit_worksheet(writer.sheets[sheet])

    if use_gcp:
        gcp_interactions_helper.write_file_to_gcp(
            excel_buffer.getvalue(), extra_info_path
        )


def process_papers(
    events=[],
    excel_path="",
    tj_pdf_base_path="",
    judges_names=[],
    tj_filter=[],
    use_gcp=False,
    include_additional_analysis=False,
    sheet_per_trial_judge=False
):
    workbook = openpyxl.Workbook()
    workbook_per_tj_dict={}
    if sheet_per_trial_judge:
        for trial_judge in judges_names:
            workbook_per_tj_dict[trial_judge] = openpyxl.Workbook()

    print(f"processing for {events} and judges {judges_names}")
    judge_errors = {}
    event_details = {}
    agg_average_deviations = {judge: {} for judge in judges_names}
    agg_all_element_df = None
    agg_all_pcs_df = None

    i = 0
    for event in events:
        pdf_path = f"{tj_pdf_base_path}{event}.pdf"
        tj_xlsx_path = f"{tj_pdf_base_path}{event}_analysis.xlsx"
        (
            event_name,
            total_errors,
            num_starts,
            allowed_errors,
            average_deviation,
            all_element_df,
            all_pcs_df,
        ) = analyze_trial_judges(
            tj_xlsx_path,
            workbook,
            pdf_path,
            judges_names,
            2,
            tj_filter,
            use_gcp=use_gcp,
            per_trial_judge_workbook_dict=workbook_per_tj_dict
        )

        # Summary statistics
        for judge in average_deviation:
            agg_average_deviations[judge][event_name] = average_deviation[judge]
        if agg_all_pcs_df is None:
            agg_all_pcs_df = all_pcs_df
        else:
            agg_all_pcs_df = pd.concat([agg_all_pcs_df, all_pcs_df])

        if agg_all_element_df is None:
            agg_all_element_df = all_element_df
        else:
            agg_all_element_df = pd.concat([agg_all_element_df, all_element_df])

        start_of_summary_rows = sum(total_errors) + 11
        event_details[event_name] = {
            "Num Starts": num_starts,
            "Allowed Errors": allowed_errors,
            "Sheet Name": judgingParsing.get_sheet_name(event_name, 2),
            "Summary Row Start": start_of_summary_rows,
        }
        for i in range(len(judges_names)):
            judge = judges_names[i]
            if len(tj_filter) > 0 and judge not in tj_filter:
                continue
            if judge not in judge_errors:
                judge_errors[judge] = {}
            judge_errors[judge][event_name] = {
                "Errors": total_errors[i],
                "Allowed Errors": allowed_errors,
                "In Excess": max(total_errors[i] - allowed_errors, 0),
                "Judge Number": i + 1,
            }

    # Sort sheets
    del workbook["Sheet"]
    workbook._sheets.sort(key=lambda ws: ws.title)
    for trial_judge in workbook_per_tj_dict:
        tj_workbook = workbook_per_tj_dict[trial_judge]
        workbook_per_tj_dict[trial_judge]._sheets.sort(key=lambda ws: ws.title)
        filtered_judge_errors ={trial_judge: judge_errors[trial_judge]}
        make_competition_summary_page(tj_workbook, "Trial Judge", event_details, filtered_judge_errors)

    make_competition_summary_page(workbook, "Trial Judge", event_details, judge_errors)

    if use_gcp:
        gcp_interactions_helper.save_gcp_workbook(workbook, excel_path)
        for trial_judge in workbook_per_tj_dict:
            path_to_use = excel_path.replace(".xlsx", f"_{trial_judge.replace(" ", "_")}.xlsx")
            gcp_interactions_helper.save_gcp_workbook(workbook_per_tj_dict[trial_judge], path_to_use)
    else:
        workbook.save(excel_path)
        for trial_judge in workbook_per_tj_dict:
            path_to_use = excel_path.replace(".xlsx", f"_{trial_judge.replace(" ", "_")}.xlsx")
            workbook_per_tj_dict[trial_judge].save(path_to_use)

    if include_additional_analysis:
        make_extra_analysis_sheet(
            excel_path,
            agg_all_element_df,
            agg_all_pcs_df,
            agg_average_deviations,
            use_gcp=use_gcp,
        )


def format_out_of_range_sheets(worksheet):
    color_scale_rule = ColorScaleRule(
        start_type="min",
        start_color="FFFFFF",  # White
        #  mid_type='percentile', mid_value=50, mid_color='7FFFD4',
        end_type="max",
        end_color="FF0000",
    )  # Red
    worksheet.conditional_formatting.add("F2:H200", color_scale_rule)
    for cell in worksheet["F"]:
        cell.number_format = "0%"
    for cell in worksheet["G"]:
        cell.number_format = "0%"
    for cell in worksheet["H"]:
        cell.number_format = "0%"


if __name__ == "__main__":
    # Specify paths for the input PDF and output Excel file
    # pdf_base_path = "/Users/rnaphtal/Documents/JudgingAnalysis/TrialJudges/"  # Update with the correct path
    excel_path = "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/TrialJudges/ORC_Anomaly_Summary_Analysis.xlsx"
    tj_pdf_base_path = (
        "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/TrialJudges/"
    )

    # GCP example
    # excel_path = "skating_orc_reports/gs://skating_orc_reports/Generated/Nats/Nats.xlsx"
    # tj_pdf_base_path = "skating_orc_reports/gs://skating_orc_reports/Generated/Nats/"

    workbook = openpyxl.Workbook()
    events = ["NPFS", "JMFS", "JMSP", "JPSP", "JPFS", "SWSP", "SPSP", "SWFS", "SMSP"]
    # events = ["JMSP"]

    judgesNames = [
        "Melanya Berggren",
        "Katie Beriau",
        "Scott Brody",
        "Waverly Huston",
        "Rhea Sy-Benedict",
        "William Tran",
        "Mary-E Wightman",
    ]
    # process_papers(
    #     use_gcp=False,
    #     events=events,
    #     excel_path=excel_path,
    #     tj_pdf_base_path=tj_pdf_base_path,
    #     judges_names=judgesNames,
    #     include_additional_analysis=True
    # )
    # print (get_judges_name_from_sheet(f"{tj_pdf_base_path}JMFS_analysis.xlsx", use_gcp=True))
    # for judge in judgesNames:
    # processPapers(events=events, excel_path=f"{tj_pdf_base_path}ORC_Anomaly_Summary_Analysis_{judge}.xlsx", tj_pdf_base_path=tj_pdf_base_path, judges_names=judgesNames, tj_filter=[judge])

    # 2024
    excel_path = "/Users/rnaphtal/Documents/JudgingAnalysis/TrialJudges/2024/ORC_Anomaly_Summary_Analysis_Melanya_and_Scott.xlsx"
    tj_pdf_base_path = "/Users/rnaphtal/Documents/JudgingAnalysis/TrialJudges/2024/"
    events = ["NPSP", "JPSP", "JPFS", "SPSP", "SWSP", "SWFS", "SMSP"]
    judgesNames = [
        "Melanya Berggren",
        "Scott Brody",
        "Shelbi Gill",
        "Joy Jin",
        "Elliot Schwartz",
        "Mary-E Wightman",
    ]
    # processPapers(events=events, excel_path=f"{tj_pdf_base_path}ORC_Anomaly_Summary_Analysis_Melanya.xlsx", tj_pdf_base_path=tj_pdf_base_path, judges_names=judgesNames, tj_filter=["Melanya Berggren"])
    # processPapers(events=events, excel_path=f"{tj_pdf_base_path}ORC_Anomaly_Summary_Analysis_Scott.xlsx", tj_pdf_base_path=tj_pdf_base_path, judges_names=judgesNames, tj_filter=["Scott Brody"])

    excel_path = "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/TrialJudges/2024_SYS/ORC_Anomaly_Summary_Analysis.xlsx"
    tj_pdf_base_path = (
        "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/TrialJudges/2024_SYS/"
    )

    events = ["12FS", "12SP", "ATFS", "ITFS", "JTFS", "JTSP", "JvTFS", "NTFS", "STSP"]
    judgesNames = [
        "Sherri Cleveland",
        "Felicia Haining-Miller",
        "Megan Jackson",
        "Stephanie Pusch",
        "Elise Requadt",
        "April Zak",
    ]
    process_papers(
        use_gcp=False,
        events=events,
        excel_path=excel_path,
        tj_pdf_base_path=tj_pdf_base_path,
        judges_names=judgesNames,
        include_additional_analysis=True,
        sheet_per_trial_judge=True
    )
