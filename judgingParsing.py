import openpyxl.worksheet
import pdfplumber
from pypdf import PdfReader
import pandas as pd
import re
import requests
from bs4 import BeautifulSoup
from collections import defaultdict

from sharedJudgingAnalysis import categorizeElement
from pyppeteer import launch
import time

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
    Color,
)
from openpyxl.worksheet.datavalidation import DataValidation
from google.cloud import storage
import gcsfs
from gcp_interactions_helper import read_file_from_gcp
import streamlit as st

USING_ISU_COMPONENT_METHOD = False


def autofit_worksheet(worksheet):
    max_length = 0
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        worksheet.column_dimensions[column].width = adjusted_width


def parse_scores(pdf_path, event_regex="", use_gcp=False):
    if use_gcp:
        pdf_bytes = read_file_from_gcp(pdf_path)
        with PdfReader(pdf_bytes) as pdf:
            return process_scores(pdf, event_regex=event_regex, use_gcp=use_gcp)
    else:
        with PdfReader(pdf_path) as pdf:
            return process_scores(pdf, event_regex=event_regex, use_gcp=use_gcp)


def process_scores(pdf, event_regex="", use_gcp=False):
    # Initialize list for storing extracted data
    elements_per_skater = {}
    pcs_per_skater = {}
    skater_details = {}
    event_name = ""
    start = time.time()
    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            return

        # if "JUDGES DETAILS PER SKATER" in text:
        text = text.replace("\xa0", "")

        # Split text into lines
        lines = text.split("\n")
        if event_name == "":
            event_name = (
                lines[2]
                .replace("/", "")
                .replace(" ", "_")
                .replace("__", "_")
                .replace("-", "_")
            )
            event_name = event_name.split("/")[0]
            event_name = event_name.split(":")[0]
            if not re.match(event_regex, event_name):
                return (None, None, None, event_name)

        current_skater = None
        has_bonus = False

        for i in range(len(lines)):
            line = lines[i]
            if line == "Bonus":
                has_bonus = "True"
            # Match skater's name section
            skater_match = match_skater(line, has_bonus)

            if current_skater is None and line == "# Executed":
                # The skater name was probably split across two lines
                possible_skater = f"{lines[i - 2]}{lines[i - 1]}"
                skater_match = match_skater(possible_skater, has_bonus)

            if skater_match:
                skater_name = skater_match.group(2).strip().replace("ﬁ", "fi")
                technical_score = skater_match.group(5)
                current_skater = skater_name
                if current_skater not in elements_per_skater:
                    elements_per_skater[current_skater] = []
                    pcs_per_skater[current_skater] = []
                    skater_details[current_skater] = technical_score
                if int(skater_match.group(1)) != len(elements_per_skater.keys()):
                    print(
                        f"Missing skater in {event_name}. Next is {current_skater}")
                    st.error(
                        f"Missing skater in {event_name}. Next is {current_skater}"
                    )
                continue

            # Match elements and judge scores
            element_match = re.match(
                r"""^
                (1?\d)                                   # Element number
                (\s)*                                    
                ([\w\+\.\*<>!]+(?:\w+)?)                 # Element name
                \s+
                (?:([F*<!>qnscuSCUex]+)\s+)?             # Optional notes
                (-?[\d\.]+x?)                            # Base value
                \s+
                (?:([F\*x]+)\s+)?                        # Optional flags
                (-?[\d\.]+)                              # GOE
                \s+
                (-?(?:-?\d\s+){3,9})                     # Judge scores
                \s*
                ([\d\.]+\s)?                             # Optional referee score
                ([\d\.]+)                                # Total score
                """,
                line,
                re.VERBOSE,
            )

            pcs_match = re.match(
                r"""^\s*
                (Timing|Presentation|Skating\sSkills|Composition|Artistic\sAppeal)  # Component
                \s+
                ([\d\.]+)                                    # Factor
                \s+
                ((?:\d{1,2}\.\d{2}\s*){3,9})                 # Per-judge scores
                ([\d\.]+)                                    # Total
                """,
                line,
                re.VERBOSE,
            )

            if current_skater and element_match:
                element_number = int(element_match.group(1))
                element_name = element_match.group(3)
                notes = element_match.group(4)
                judge_scores_raw = element_match.group(8)
                total_score = float(element_match.group(10))

                # Clean up element name for SEQ< or SEQ<<
                if element_name.endswith("SEQ<"):
                    element_name = element_name[:-1]
                elif element_name.endswith("SEQ<<"):
                    element_name = element_name[:-2]

                judges_scores = list(map(float, judge_scores_raw.split()))

                elements_per_skater[current_skater].append({
                    "Element": element_name,
                    "Scores": judges_scores,
                    "Value": total_score,
                    "Notes": notes,
                    "Number": element_number,
                })

            elif current_skater and pcs_match:
                component = pcs_match.group(1)
                raw_scores_with_spaces = pcs_match.group(3)
                total_score = float(pcs_match.group(4))

                # Attempt to detect missing position
                score_tokens = raw_scores_with_spaces.split(" ")
                try:
                    possible_missing_position = score_tokens.index("")
                except ValueError:
                    possible_missing_position = None

                # Reconstruct scores by joining chars until each decimal completes
                no_spaces = raw_scores_with_spaces.replace(" ", "")
                scores = []
                i = 0
                current_score = ""

                while i < len(no_spaces):
                    if no_spaces[i] == ".":
                        current_score += no_spaces[i:i+3]
                        i += 3
                        scores.append(current_score)
                        current_score = ""
                    else:
                        current_score += no_spaces[i]
                        i += 1

                judges_scores = list(map(float, scores))

                pcs_per_skater[current_skater].append({
                    "Component": component,
                    "Scores": judges_scores,
                    "Possible Missing Position": possible_missing_position
                })
            if current_skater == None and (pcs_match or element_match):
                print(
                    f"Element or pcs found without skater. Currently {len(skater_details.keys())} skaters found. Event: {event_name}"
                )
                st.error(
                    f"Element or pcs found without skater. Currently {len(skater_details.keys())} skaters found. Event: {event_name}"
                )

    for skater in elements_per_skater:
        foundElements = round(sum([x["Value"]
                              for x in elements_per_skater[skater]]), 2)
        expected = float(skater_details[skater])
        if foundElements != expected:
            print(
                f"Elements for skater {skater} do not match. Expected TES:{expected}, Sum of elements:{foundElements}"
            )
            st.error(
                f"Elements for skater {skater} do not match. Expected TES:{expected}, Sum of elements:{foundElements}"
            )
        pcs = pcs_per_skater[skater]
        if len(pcs) < 3:
            print(f"Components missing for skater {skater} {event_name}")
            st.error(f"Components missing for skater {skater} {event_name}")

    return (elements_per_skater, pcs_per_skater, skater_details, event_name)


def extract_skater_element_sections(soup):
    tables = soup.find_all("table")
    pairs = []
    i = 0
    while i < len(tables):
        tbl = tables[i]
        if "sum" in (tbl.get("class") or []):
            row = tbl.find("tbody").find("tr")

            rank = row.find_all(class_="rank")[0].get_text(
                strip=True) if row else None
            skater_title = row.find_all(class_="name")[0].get_text(
                strip=True) if row else None
            skater_name = skater_title.split(",")[0]
            skater_total_element_score = row.find_all(
                class_="totElm")[0].get_text(strip=True) if row else None
            skater = {"name": skater_name,
                      "rank": rank,
                      "element_score": skater_total_element_score}

            elm = None
            for j in range(i + 1, len(tables)):
                if "elm" in (tables[j].get("class") or []):
                    elm = tables[j]
                    break
            if skater and elm:
                pairs.append((skater, elm))
                i = j
        i += 1
    if not pairs:
        raise ValueError("No skater data found on the segment page.")
    return pairs


def process_scores_html(soup, event_regex="", use_gcp=False):
    # Initialize list for storing extracted data
    elements_per_skater = defaultdict(list)
    pcs_per_skater = defaultdict(list)
    skater_details = {}
    event_name = soup.find_all(class_="catseg")[0].text.replace("/", "")
    event_name = event_name.replace(
        "/", "").replace(" ", "_").replace("__", "_").replace("-", "_")

    event_name = event_name.split("/")[0]
    event_name = event_name.split(":")[0]

    skater_element_pairs = extract_skater_element_sections(soup)

    for (skater_info, element_section) in skater_element_pairs:
        current_skater = skater_info["name"]
        skater_details[current_skater] = skater_info
        if int(skater_info["rank"]) != len(skater_details):
            raise ValueError(
                f"Skater {skater_info["name"]} has incorrect rank. Expected {len(skater_details)} got {skater_info["rank"]}")

        rows = element_section.find_all("tr")

        for i in range(1, len(rows)):
            row = rows[i]
            tds = row.find_all("td")

            # process elements
            if "class" in tds[0].attrs and "num" in tds[0].attrs["class"]:
                element_number = int(row.find_all(class_='num')[0].text)
                element_name = row.find_all(class_='elem')[0].text
                element_info = row.find_all(class_='info')[0].text
                element_bv = float(row.find_all(class_='bv')[0].text)
                scores = [td.text.replace(u'\xa0', "")
                          for td in row.find_all(class_="jud")]
                if any(v.strip() == "-" for v in scores):
                    continue
                scores = [int(x) if x.strip() != "" else None for x in scores]
                element_total_value = float(row.find_all(class_='psv')[0].text)

                elements_per_skater[current_skater].append({
                    "Element": element_name,
                    "Scores": scores,
                    "Base Value": element_bv,
                    "Value": element_total_value,
                    "Notes": element_info,
                    "Number": element_number,
                })
            # Process PCS
            elif "class" in tds[1].attrs and "cn" in tds[1].attrs["class"]:
                component_name = row.find_all(class_='cn')[0].text
                scores = [td.text.replace(u'\xa0', "")
                          for td in row.find_all(class_="cjud")]
                scores = [float(x) if x.strip() !=
                          "" else None for x in scores]

                pcs_per_skater[current_skater].append({
                    "Component": component_name,
                    "Scores": scores,
                })

    for skater in elements_per_skater:
        foundElements = round(sum([x["Value"]
                              for x in elements_per_skater[skater]]), 2)
        expected = float(skater_details[skater]["element_score"])
        if foundElements != expected:
            print(
                f"Elements for skater {skater} do not match. Expected TES:{expected}, Sum of elements:{foundElements}"
            )
            st.error(
                f"Elements for skater {skater} do not match. Expected TES:{expected}, Sum of elements:{foundElements}"
            )
        pcs = pcs_per_skater[skater]
        if len(pcs) < 3:
            print(f"Components missing for skater {skater} {event_name}")
            st.error(f"Components missing for skater {skater} {event_name}")

    return (elements_per_skater, pcs_per_skater, skater_details, event_name)


def create_all_element_dict(judges, elements_per_skater, event_name):
    all_elements = []
    for skater in elements_per_skater:
        for element in elements_per_skater[skater]:
            all_scores = element["Scores"]
            filtered_scores = [
                score for score in all_scores if score is not None]
            avg = sum(filtered_scores) / len(filtered_scores)
            judgeNumber = 1
            if len(all_scores) < len(judges):
                raise ValueError(
                    f"Missing components in {event_name} for {skater}, element: {element}")
            for judge in judges:
                judge_score = all_scores[judgeNumber - 1]
                if judge_score is None:
                    print(
                        f"Missing elements for skater {skater} judge {judgeNumber}")
                    continue

                deviation = judge_score - avg
                thrown_out = is_score_thrown_out(judge_score, all_scores)
                high = judge_score > avg
                element_name_no_level = element["Element"]
                if element_name_no_level[-1].isdigit():
                    element_name_no_level = element_name_no_level[:-1]

                all_elements.append(
                    {
                        "Skater": skater,
                        "Event": event_name,
                        "Element": element["Element"],
                        "Element Type": categorizeElement(element_name_no_level),
                        "Panel Average": avg,
                        "Judge Name": judge,
                        "Judge Number": judgeNumber,
                        "Score": judge_score,
                        "Deviation": deviation,
                        "Thrown out": thrown_out,
                        "High": high,
                    }
                )
                judgeNumber += 1
    return all_elements


def create_all_pcs_dict(judges, pcs_per_skater, event_name):
    all_pcs = []
    for skater in pcs_per_skater:
        for pcs_mark in pcs_per_skater[skater]:
            all_scores = pcs_mark["Scores"]
            if len(all_scores) < len(judges):
                missing_position = pcs_mark["Possible Missing Position"]
                all_scores.insert(missing_position, 0)

            filtered_scores = [
                score for score in all_scores if score is not None]
            avg = sum(filtered_scores) / len(filtered_scores)
            judgeNumber = 1
            for judge in judges:
                judge_score = all_scores[judgeNumber - 1]
                if judge_score is None:
                    print(f"Missing components for skater {skater}")
                    continue
                deviation = judge_score - avg
                thrown_out = is_score_thrown_out(judge_score, all_scores)
                high = judge_score > avg
                all_pcs.append(
                    {
                        "Skater": skater,
                        "Event": event_name,
                        "Component": pcs_mark["Component"],
                        "Panel Average": avg,
                        "Judge Name": judge,
                        "Judge Number": judgeNumber,
                        "Score": judge_score,
                        "Deviation": deviation,
                        "Thrown out": thrown_out,
                        "High": high,
                    }
                )
                judgeNumber += 1
    return all_pcs


def is_score_thrown_out(score, all_scores):
    filtered_scores = [score for score in all_scores if score is not None]
    max_panel = max(filtered_scores)
    min_panel = min(filtered_scores)

    # It should not count if at least 3 judges agree.
    count_same = sum([1 for s in all_scores if s == score])
    if count_same >= 3:
        return False
    return score == min_panel or score == max_panel


def extract_judge_scores(
    workbook,
    pdf_path,
    base_excel_path,
    judges,
    pdf_number,
    use_html=True,
    url="",
    event_regex="",
    only_rule_errors=False,
    use_gcp=False,
    create_thrown_out_analysis=False,
    judge_filter=""
):
    if use_html:
        page_contents = get_page_contents(url)
        soup = BeautifulSoup(page_contents, "html.parser")
        (elements_per_skater, pcs_per_skater, skater_details, event_name) = process_scores_html(
            soup=soup, event_regex=event_regex, use_gcp=use_gcp
        )
    else:
        (elements_per_skater, pcs_per_skater, skater_details, event_name) = parse_scores(
            pdf_path, event_regex, use_gcp=use_gcp
        )
    if not re.match(event_regex, event_name):
        return (event_name, None, None, None, [], [], [])

    element_errors = []
    element_deviations = []
    pcs_errors = []
    if (
        "Women" in event_name
        or "Men" in event_name
        or "Boys" in event_name
        or "Girls" in event_name
    ):
        element_errors = findSinglesElementErrors(
            elements_per_skater, judges, event_name, judge_filter=judge_filter
        )
    elif "Pairs" in event_name:
        element_errors = findPairsElementErrors(
            elements_per_skater, judges, event_name, judge_filter=judge_filter)

    if not only_rule_errors:
        element_deviations = findElementDeviations(
            elements_per_skater, judges, judge_filter=judge_filter)
        pcs_errors = findPCSDeviations(
            pcs_per_skater, judges, judge_filter=judge_filter)
    total_errors = count_total_errors_per_judge(
        judges, element_errors, element_deviations, pcs_errors
    )

    formatted_event_name = event_name.replace("/", "")
    # event_name = (
    #             lines[2]
    #             .replace("/", "")
    #             .replace(" ", "_")
    #             .replace("__", "_")
    #             .replace("-", "_")
    #         )
    #         event_name = event_name.split("/")[0]
    #         event_name = event_name.split(":")[0]
    printToExcel(
        workbook,
        event_name,
        judges,
        element_errors,
        element_deviations,
        pcs_errors,
        pdf_number,
    )

    all_element_dict = {}
    all_pcs_dict = {}
    if create_thrown_out_analysis:
        all_element_dict = create_all_element_dict(
            judges, elements_per_skater, event_name
        )
        all_pcs_dict = create_all_pcs_dict(judges, pcs_per_skater, event_name)
    return (
        event_name,
        total_errors,
        len(skater_details),
        get_allowed_errors(len(skater_details)),
        element_errors,
        all_element_dict,
        all_pcs_dict,
    )


def match_skater(line, has_bonus):
    if has_bonus:
        return re.match(
            r"^(\d+)\s+([A-Za-z1-9!\.\(\)\/\-\s'ﬁ]+[A-Za-z!\.\(\)\/\-\s'ﬁ]+),?([A-Za-z1-9\&\(\)\-\.\s]+[A-Za-z\&\(\)\-\.\s]+)?\s?([\d]{1,3}\.[\d\.]{2})\s?([\d]{1,3}\.[\d\.]{2})\s?([\d]{1,3}\.[\d\.]{2})\s?([\d]{1,3}\.[\d\.]{2})\s?([\d]{1,3}\.[\d\.]{2})$",
            line,
        )
    else:
        return re.match(
            r"^(\d+)\s+([A-Za-z1-9!\.\(\)\/\-\s'ﬁ]+[A-Za-z!\.\(\)\/\-\s'ﬁ]+),?([A-Za-z1-9\&\(\)\-\.\s]+[A-Za-z\&\(\)\-\.\s]+)?\s?([\d]{1,3}\.[\d\.]{2})\s?([\d]{1,3}\.[\d\.]{2})\s?([\d]{1,3}\.[\d\.]{2})\s?([\d]{1,3}\.[\d\.]{2})?\s?([\d]{1,3}\.[\d\.]{2})$",
            line,
        )


def get_allowed_errors(num_skaters: int):
    if num_skaters <= 10:
        return 1
    elif num_skaters <= 20:
        return 2
    return 3


def findSinglesElementErrors(skater_scores, judges, event_name, judge_filter=""):
    errors = []
    for skater in skater_scores:
        for element in skater_scores[skater]:
            element_name = element["Element"]
            notes = element["Notes"]
            allScores = element["Scores"]

            judgeNumber = 1
            for judgeNumber in range(1, len(allScores) + 1):
                
                if allScores[judgeNumber - 1] is None:
                    # This is a missing score or judge that isn't included
                    if judgeNumber < len(judges)+1:
                        print(
                            f"Missing elements for skater {skater} judge {judgeNumber}")
                    continue
                if len(judge_filter) > 0 and judge_filter != judges[judgeNumber-1]:
                    continue
                # Must be -5 if  it is a short and there is a +COMBO or *
                if (
                    "Short" in event_name
                    and ("COMBO" in element_name or "*" in element_name)
                    and allScores[judgeNumber - 1] > -5
                ):
                    errors.append(
                        makeRuleError(
                            skater,
                            element,
                            judgeNumber,
                            judges,
                            allScores,
                            "Short Program NAR"
                        )
                    )
                    continue

                max_goe = 5
                number_downs = element_name.count("<<")
                number_unders = element_name.count("<") - (number_downs * 2)
                number_attention = element_name.count("!")
                # If contains certain errors cannot start above a 2
                if (
                    notes is not None
                    and ("F" in notes or "e" in notes)
                    or "<<" in element_name
                    or (number_unders + number_attention) >= 2
                ):
                    max_goe = 2
                # Falls must subtract 5
                if notes is not None and "F" in notes:
                    max_goe = max_goe - 5
                # e must subtract 2
                if notes is not None and "e" in notes:
                    max_goe -= 2
                # << must subtract 3
                max_goe -= 3 * number_downs
                # Subtract 2 for <
                max_goe -= 2 * number_unders
                # Subtract 2 for q
                if notes is not None and "q" in notes:
                    max_goe -= 2 * element_name.count("q")
                # Subtract 1 for attention
                max_goe -= 1 * number_attention

                if allScores[judgeNumber - 1] > max(max_goe, -5):
                    errors.append(
                        makeRuleError(
                            skater,
                            element,
                            judgeNumber,
                            judges,
                            allScores,
                            f"Max with errors is {max(max_goe, -5)}",
                        )
                    )

    return errors


def findPairsElementErrors(skater_scores, judges, event_name, judge_filter=""):
    return findSinglesElementErrors(skater_scores, judges, event_name, judge_filter=judge_filter)


def makeRuleError(skater, element, judgeNumber, judges, allScores, description):
    element_name = f"{element['Element']} {element['Notes']}"
    if element["Notes"] is None:
        element_name = element["Element"]
    return {
        "Skater": skater,
        "Element": element_name,
        "Judge Number": judgeNumber,
        "Judge Name": judges[judgeNumber - 1],
        "Judge Score": allScores[judgeNumber - 1],
        "Panel Average": None,
        "Deviation": "Rule Error",
        "Description": description,
    }


def findElementDeviations(skater_scores, judges, judge_filter=""):
    errors = []
    for skater in skater_scores:
        for element in skater_scores[skater]:
            allScores = element["Scores"]
            filtered_scores = [
                score for score in allScores if score is not None]
            avg = sum(filtered_scores) / len(filtered_scores)
            judgeNumber = 1
            for judgeNumber in range(1, len(allScores) + 1):
                if allScores[judgeNumber - 1] is None:
                    if judgeNumber < len(judges)+1:
                        print(
                            f"Missing elements for skater {skater} judge {judgeNumber}")
                    continue
                if len(judge_filter) > 0 and judge_filter != judges[judgeNumber-1]:
                    continue
                deviation = allScores[judgeNumber - 1] - avg
                if abs(deviation) >= 2:
                    # print (f"Deviation found for judge {judgeNumber} on skater {skater}, element {element["Element"]}")
                    errors.append(
                        {
                            "Skater": skater,
                            "Element": element["Element"],
                            "Judge Number": judgeNumber,
                            "Judge Name": judges[judgeNumber - 1],
                            "Judge Score": allScores[judgeNumber - 1],
                            "Panel Average": avg,
                            "Deviation": deviation,
                            "Type": "Deviation",
                        }
                    )
    return errors


def findPCSDeviations(skater_scores, judges, judge_filter=""):
    errors = []
    for skater in skater_scores:
        deviation_points = [float(0)] * (len(judges) + 1)
        if len(skater_scores[skater]) == 0:
            continue
        for component in skater_scores[skater]:
            allScores = component["Scores"]
            filtered_scores = [
                score for score in allScores if score is not None]
            avg = sum(filtered_scores) / len(filtered_scores)
            for judgeNumber in range(1, len(allScores) + 1):
                if allScores[judgeNumber - 1] is None:
                    if judgeNumber < len(judges)+1:
                        print(
                            f"Missing component for skater {skater} judge {judgeNumber}")
                    continue
                if len(judge_filter) > 0 and judge_filter != judges[judgeNumber-1]:
                    continue
                deviation = allScores[judgeNumber - 1] - avg
                if not USING_ISU_COMPONENT_METHOD and abs(deviation) >= 1.5:
                    errors.append(
                        {
                            "Skater": skater,
                            "Judge Number": judgeNumber,
                            "Judge Name": judges[judgeNumber - 1],
                            "Judge Score": allScores[judgeNumber - 1],
                            "Deviation": deviation,
                            "Component": component["Component"],
                            "Type": "Deviation",
                        }
                    )
                deviation_points[judgeNumber] = (
                    deviation_points[judgeNumber] + deviation
                )

        for judgeNumber in range(1, len(allScores) + 1):
            if USING_ISU_COMPONENT_METHOD and deviation_points[judgeNumber] > 4.5:
                # Add errors here if using ISU method
                errors.append(
                    {
                        "Skater": skater,
                        "Judge Number": judgeNumber,
                        "Judge Name": judges[judgeNumber - 1],
                        "Judge Score": "",
                        "Deviation": deviation_points[judgeNumber],
                    }
                )
    return errors


def count_total_errors_per_judge(
    judges, element_errors, element_deviations, pcs_deviations
) -> list:
    errors_list = [0] * len(judges)
    for error in element_errors:
        judgeNumber = error["Judge Number"]
        errors_list[judgeNumber - 1] = errors_list[judgeNumber - 1] + 1
    for error in element_deviations:
        judgeNumber = error["Judge Number"]
        errors_list[judgeNumber - 1] = errors_list[judgeNumber - 1] + 1
    for error in pcs_deviations:
        judgeNumber = error["Judge Number"]
        errors_list[judgeNumber - 1] = errors_list[judgeNumber - 1] + 1
    return errors_list


def get_sheet_name(event_name, pdf_number):
    sheet_name = event_name
    with_el_match = re.match(r"(\d{1,3})_(\d{1,3}_)?(.*)", sheet_name)
    sheet_name = sheet_name.replace("(", "").replace(")", "").replace("&", "")
    if with_el_match:
        sheet_name = with_el_match.group(3)
    if len(sheet_name) >= 28:
        sheet_name = f"{sheet_name[0: min(len(sheet_name), 26)]}{pdf_number}"
    return sheet_name


def printToExcel(
    workbook,
    event_name,
    judges,
    element_errors,
    element_deviations,
    pcs_deviations,
    pdf_number,
):
    sheet_name = get_sheet_name(event_name, pdf_number)

    bold = Font(bold=True)
    gray = PatternFill("solid", fgColor="C0C0C0")
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    sheet = workbook.create_sheet(sheet_name, 0)
    sheet.cell(1, 1, value=event_name.replace("_", " ")).font = bold
    yes_no = DataValidation(
        type="list", formula1='"YES,NO"', showDropDown=False, allow_blank=True
    )
    yes_no.showInputMessage = True
    yes_no.showErrorMessage = True
    if len(element_deviations) + len(element_errors) + len(pcs_deviations) > 0:
        sheet.add_data_validation(yes_no)
    # Headers
    sheet.cell(4, 1, value="Judge").font = bold
    sheet.cell(4, 2, value="Judge Score").font = bold
    sheet.cell(4, 3, value="Deviation From Panel Average").font = bold
    sheet.cell(4, 4, value="Skater(s)/Couple(s)").font = bold
    sheet.cell(4, 5, value="Element Name").font = bold
    sheet.cell(4, 6, value="ORC Comments").font = bold
    sheet.cell(4, 7, value="ORC Error?").font = bold
    sheet.cell(5, 1, value="A. RANGES OF GOE").font = bold

    current_row = 6
    for error in element_errors:
        sheet.cell(
            current_row, 1, value=f"J{error['Judge Number']}- {error['Judge Name']}"
        )
        sheet.cell(current_row, 2, value=error["Judge Score"])
        sheet.cell(current_row, 3, value=error["Description"])
        sheet.cell(current_row, 4, value=error["Skater"])
        sheet.cell(current_row, 5, value=error["Element"])
        sheet.cell(current_row, 7).value = "YES"
        yes_no.add(sheet.cell(current_row, 7))
        sheet.cell(current_row, 6).fill = gray
        sheet.cell(current_row, 6).border = thin_border
        sheet.cell(current_row, 7).fill = gray
        sheet.cell(current_row, 7).border = thin_border
        sheet.cell(current_row, 6).alignment = Alignment(wrap_text=True)
        current_row = current_row + 1

    for error in element_deviations:
        sheet.cell(
            current_row, 1, value=f"J{error['Judge Number']}- {error['Judge Name']}"
        )
        sheet.cell(current_row, 2, value=error["Judge Score"])
        sheet.cell(current_row, 3, value=error["Deviation"])
        sheet.cell(current_row, 4, value=error["Skater"])
        sheet.cell(current_row, 5, value=error["Element"])
        yes_no.add(sheet.cell(current_row, 7))
        sheet.cell(current_row, 6).fill = gray
        sheet.cell(current_row, 6).border = thin_border
        sheet.cell(current_row, 7).fill = gray
        sheet.cell(current_row, 7).border = thin_border
        sheet.cell(current_row, 6).alignment = Alignment(wrap_text=True)
        current_row = current_row + 1

    current_row += 1
    sheet.cell(current_row, 1).value = "B. RANGES OF PROGRAM COMPONENTS"
    sheet.cell(current_row, 1).font = bold
    current_row += 1
    for error in pcs_deviations:
        sheet.cell(
            current_row,
            1,
            value=str(f"J{error['Judge Number']}- {error['Judge Name']}"),
        )
        sheet.cell(current_row, 2, value=error["Judge Score"])
        sheet.cell(current_row, 3, value=str(error["Deviation"]))
        sheet.cell(current_row, 4, value=error["Skater"])
        sheet.cell(current_row, 5, value=error["Component"])
        yes_no.add(sheet.cell(current_row, 7))
        sheet.cell(current_row, 6).fill = gray
        sheet.cell(current_row, 6).border = thin_border
        sheet.cell(current_row, 6).alignment = Alignment(wrap_text=True)
        sheet.cell(current_row, 7).fill = gray
        sheet.cell(current_row, 7).border = thin_border

        current_row = current_row + 1
    cell_end_errors_section = current_row

    current_row += 2
    sheet.cell(current_row, 1, value="Judge").font = bold
    sheet.cell(current_row, 1).font = Font(bold=True)
    sheet.cell(current_row, 2, value="# of Anomalies").font = bold
    sheet.cell(current_row, 3, value="ORC Recognized Error").font = bold
    current_row += 1

    for i in range(len(judges)):
        sheet.cell(current_row, 1, value=f"J{i + 1}- {judges[i]}")
        sheet.cell(
            current_row,
            2,
            value=f"=COUNTIF(A$6:A${cell_end_errors_section},A{current_row})",
        )
        sheet.cell(
            current_row,
            3,
            value=f'=COUNTIFS(A$6:A${cell_end_errors_section},A{current_row},G$6:G${cell_end_errors_section},"YES")',
        )
        current_row += 1

    autofit_worksheet(sheet)
    sheet.column_dimensions["F"].width = 35
    # print (f"Processed {event_name}")


def get_page_contents(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36"
    }

    page = requests.get(url, headers=headers)

    if page.status_code == 200:
        return page.text

    return None


if __name__ == "__main__":
    # Specify paths for the input PDF and output Excel file
    # Update with the correct path
    pdf_path = "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/Easterns/Results/1.pdf"
    excel_path = "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/ISU/"
    tj_pdf_path = "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/ISU/FC.xlsx"

    workbook = openpyxl.Workbook()
    extract_judge_scores(
        workbook,
        pdf_path,
        excel_path,
        [
            "Name1",
            "Name2",
            "Name3",
            "Name4",
            "Name5",
            "Name6",
            "Name7",
            # "Name8",
            # "Name9",
        ],
        2,
        use_html=True,
        url='https://ijs.usfigureskating.org/leaderboard/results/2025/36275/SEGM042.html',
        create_thrown_out_analysis=True
    )
    excel_path = f"{excel_path}DeviationsReport2.xlsx"
    workbook.save(excel_path)
    # page_contents = get_page_contents("https://ijs.usfigureskating.org/leaderboard/results/2025/35783/SEGM005.html")
    # soup = BeautifulSoup(page_contents, "html.parser")
    # print(process_scores_html(soup))
