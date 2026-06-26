"""Tests for National SP judge analysis Excel export."""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

_REPO = Path(__file__).resolve().parents[1]
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

from scripts.national_sp_judge_analysis_xlsx import (  # noqa: E402
    ANALYSIS_FREEZE_PANES,
    LOOKUP_SHEET,
    RAW_SHEET,
    ANALYSIS_SHEET,
    _anomaly_rules,
    _sectionals_activity_flag,
    analysis_row_order,
    marking_score_rating,
    performance_overall,
    write_national_sp_judge_analysis_xlsx,
)
from scripts.report_national_singles_pairs_judge_errors import (  # noqa: E402
    ACTIVITY_WINDOW_SEASONS,
    REPORT_DISCIPLINE_CONFIGS,
    SYNCHRO_JUNIOR_SENIOR_MIN_TEAM_COUNT,
    SYNCHRO_JUNIOR_SENIOR_SEGMENT_COUNT_HEADER,
    SYNCHRO_JUNIOR_SENIOR_ACTIVITY_LABEL,
    _deviation_benchmark_kwargs,
    _official_candidate_judge_ids,
    _pool_judge_stats,
    _segment_counts_for_judge_ids,
    _season_years_for_listing,
)
from scripts.national_judge_report_thresholds import (  # noqa: E402
    ICE_DANCE_THRESHOLDS,
    SYNCHRO_THRESHOLDS,
)
from element_deviation_ranking import ELEMENT_RANKING_LEVEL_FILTER_ALL  # noqa: E402
from officials_competition_types import (  # noqa: E402
    COMPETITION_SCOPE_ALL,
    COMPETITION_SCOPE_QUALIFYING,
    COMPETITION_SCOPE_SPD_SECTIONALS,
    COMPETITION_SCOPE_SYS_SECTIONALS,
    SPD_SECTIONAL_TYPE_IDS,
    SYS_SECTIONAL_TYPE_IDS,
)


def test_season_years_for_listing_2627_default_window():
    assert _season_years_for_listing(2627, 3) == ["2526", "2425", "2324"]


def test_discipline_configs():
    sp = REPORT_DISCIPLINE_CONFIGS["singles_pairs"]
    dance = REPORT_DISCIPLINE_CONFIGS["ice_dance"]
    synchro = REPORT_DISCIPLINE_CONFIGS["synchro"]

    assert sp.include_rule_errors is True
    assert dance.include_rule_errors is False
    assert synchro.include_rule_errors is False

    assert sp.championships_type_ids == (4,)
    assert dance.championships_type_ids == (4,)
    assert synchro.championships_type_ids == (8,)

    assert dance.segment_discipline_type_ids == (3,)
    assert synchro.qualifying_competition_group == "synchronized"

    assert sp.activity_competition_scope == COMPETITION_SCOPE_QUALIFYING
    assert dance.activity_competition_scope == COMPETITION_SCOPE_ALL
    assert synchro.activity_competition_scope == COMPETITION_SCOPE_ALL
    assert dance.thresholds is ICE_DANCE_THRESHOLDS
    assert synchro.thresholds is SYNCHRO_THRESHOLDS
    assert dance.thresholds.anomaly_pct_fair == 1.1
    assert synchro.thresholds.anomaly_pct_fair == SYNCHRO_THRESHOLDS.anomaly_pct_fair
    assert "all activity" in dance.performance_block_header
    assert dance.recent_period_header == "Last 3 years (all activity)"
    assert synchro.junior_senior_min_team_count == SYNCHRO_JUNIOR_SENIOR_MIN_TEAM_COUNT
    assert (
        synchro.junior_senior_segment_count_header
        == SYNCHRO_JUNIOR_SENIOR_SEGMENT_COUNT_HEADER
    )
    assert synchro.junior_senior_activity_label == SYNCHRO_JUNIOR_SENIOR_ACTIVITY_LABEL
    assert sp.junior_senior_min_team_count is None
    assert synchro.champs_performance_segment_levels is not None
    assert synchro.champs_segment_level_preset == "junior_senior"
    assert sp.champs_performance_segment_levels is None
    assert synchro.thresholds.element_marking_score_fair == 1.1

    assert sp.sectionals_type_ids == tuple(sorted(SPD_SECTIONAL_TYPE_IDS))
    assert dance.sectionals_type_ids == tuple(sorted(SPD_SECTIONAL_TYPE_IDS))
    assert synchro.sectionals_type_ids == tuple(sorted(SYS_SECTIONAL_TYPE_IDS))
    assert sp.sectionals_competition_scope == COMPETITION_SCOPE_SPD_SECTIONALS
    assert synchro.sectionals_competition_scope == COMPETITION_SCOPE_SYS_SECTIONALS
    assert synchro.sectionals_performance_segment_levels is not None
    assert synchro.sectionals_segment_level_preset == "junior_senior"
    assert ACTIVITY_WINDOW_SEASONS == 2


def test_marking_score_rating_thresholds():
    assert marking_score_rating(0.05) == "N/A"
    assert marking_score_rating(0.5) == "Good"
    assert marking_score_rating(1.0) == "Fair"
    assert marking_score_rating(1.1) == "Fair"
    assert marking_score_rating(1.5) == "Poor"
    assert marking_score_rating(1.05, fair_at=1.1) == "Good"
    assert marking_score_rating(1.1, fair_at=1.1) == "Fair"


def test_sectionals_activity_flag_uses_three_year_window():
    assert _sectionals_activity_flag(2024, min_calendar_year=2023) == ""
    assert _sectionals_activity_flag(2022, min_calendar_year=2023) == "Low"
    assert _sectionals_activity_flag(None, min_calendar_year=2023) == "Low"
    assert _sectionals_activity_flag(None, min_calendar_year=2023, international_flag="X") == ""


def test_performance_overall_any_na():
    assert performance_overall("Poor", "N/A", "Good", "Fair") == "N/A"


def test_analysis_row_order_puts_poor_first():
    df = pd.DataFrame(
        [
            {
                "directory_name": "Good Judge",
                "total_comps_in_role_2yr": 20,
                "activity_competition_count": 10,
                "activity_segment_count": 50,
                "activity_junior_senior_segment_count": 30,
                "total_rule_errors": 0,
                "anomaly_rate_pct": 0.5,
                "pcs_marking_score": 0.5,
                "element_marking_score": 0.5,
                "champs_competition_count": 0,
            },
            {
                "directory_name": "Poor Judge",
                "total_comps_in_role_2yr": 20,
                "activity_competition_count": 10,
                "activity_segment_count": 50,
                "activity_junior_senior_segment_count": 30,
                "total_rule_errors": 5,
                "anomaly_rate_pct": 2.5,
                "pcs_marking_score": 1.5,
                "element_marking_score": 1.5,
                "champs_competition_count": 0,
            },
        ]
    )
    ordered = analysis_row_order(df)
    assert ordered.iloc[0]["directory_name"] == "Poor Judge"


def test_anomaly_rule_formulas_use_thresholds():
    from scripts.national_judge_report_thresholds import SINGLES_PAIRS_THRESHOLDS

    formulas = [
        rule.formula[0]
        for rule in _anomaly_rules("N4", thresholds=SINGLES_PAIRS_THRESHOLDS)
    ]
    poor = SINGLES_PAIRS_THRESHOLDS.anomaly_pct_poor
    fair = SINGLES_PAIRS_THRESHOLDS.anomaly_pct_fair
    assert any(f">={poor}" in f for f in formulas)
    assert any(f">={fair}" in f and f"<{poor}" in f for f in formulas)
    assert any(f"<{fair}" in f for f in formulas)


def test_write_workbook_layout(tmp_path: Path):
    df = pd.DataFrame(
        [
            {
                "directory_name": "Example Judge",
                "international_judge": True,
                "mbr_number": "123456",
                "us_champs_senior_availability": "Available",
                "appointment_year": 2010,
                "last_champs_in_role": 2020,
                "last_sectionals_in_role": 2024,
                "total_comps_in_role_2yr": 12,
                "activity_competition_count": 4,
                "activity_segment_count": 25,
                "activity_junior_senior_segment_count": 15,
                "pcs_marking_score": 1.05,
                "element_marking_score": 0.95,
                "competition_count": 4,
                "segment_count": 25,
                "junior_senior_segment_count": 15,
                "total_rule_errors": 2,
                "anomaly_rate_pct": 1.2,
                "sectionals_pcs_marking_score": None,
                "sectionals_element_marking_score": None,
                "sectionals_competition_count": 0,
                "sectionals_segment_count": 0,
                "sectionals_junior_senior_segment_count": 0,
                "sectionals_total_rule_errors": 0,
                "sectionals_anomaly_rate_pct": None,
                "champs_pcs_marking_score": None,
                "champs_element_marking_score": None,
                "champs_competition_count": 0,
                "champs_segment_count": 0,
                "champs_junior_senior_segment_count": 0,
                "champs_total_rule_errors": 0,
                "champs_anomaly_rate_pct": None,
            }
        ]
    )
    out = tmp_path / "report.xlsx"
    write_national_sp_judge_analysis_xlsx(df, out, sectionals_activity_min_year=2023)

    wb = load_workbook(out, data_only=False)
    assert wb.sheetnames == [ANALYSIS_SHEET, RAW_SHEET, LOOKUP_SHEET]

    ws = wb[ANALYSIS_SHEET]
    assert ws["A3"].value == "Name"
    assert ws["H2"].value == "Last 3 years"
    assert ws["G3"].value == "Last Sectionals in Role"
    assert ws["G4"].value == 2024
    assert ws["A4"].value == "Example Judge"
    assert ws["B4"].value == "X"
    assert ws["C4"].value == "123456"
    assert ws["N4"].value == f"=VLOOKUP(O4,'{LOOKUP_SHEET}'!A$2:B$5,2,TRUE)"
    assert ws["O4"].value == 0.95
    assert ws["Q4"].value == 1.05
    assert ws["I3"].value == "# Competitions"
    assert ws["N3"].value == "Element Dev Score"
    assert ws["O3"].value == "Element Marking Score"
    assert ws["P3"].value == "PCS Dev Score"
    assert ws["Q3"].value == "PCS Marking Score"
    assert ws["H3"].value == "Total Comps (2 years) in Role"
    assert ws["AV4"].value.startswith("=IF(COUNTIF(AR4:AU4")
    assert ws["AQ4"].value.startswith("=IF(COUNTIF(AM4:AP4")
    assert "@" not in str(ws["AQ4"].value)
    for visible_col in ("I", "J", "O", "Q", "R", "AA"):
        assert ws.column_dimensions[visible_col].hidden is not True
    assert ws.column_dimensions["AJ"].hidden is True
    assert ws.column_dimensions["AK"].hidden is True
    assert ws.column_dimensions["AL"].hidden is True
    assert ws["AM3"].value == "Total Activity"
    assert ws["AN3"].value == "Qualifying Activity"
    assert ws["AP3"].value == "Sectionals Activity"
    assert ws["R2"].value == "Sectionals (since 2018 for GOEs and 2022 for PCS)"
    assert ws["AW2"].value == "Sectionals Performance Analysis"
    assert ws.freeze_panes == ANALYSIS_FREEZE_PANES
    assert ws.row_dimensions[3].height == 51
    assert ws["A3"].alignment.wrap_text is True
    assert ws.column_dimensions["I"].width == 13.83
    assert ws["H4"].border.left.style == "thin"
    assert ws["AM4"].border.left.style == "thin"
    cf_ranges = {str(rng) for rng in ws.conditional_formatting._cf_rules}
    assert any("AQ" in r for r in cf_ranges)
    assert any("AV" in r for r in cf_ranges)
    assert any("BA" in r for r in cf_ranges)
    assert any("BF" in r for r in cf_ranges)
    assert any("I" in r for r in cf_ranges)
    assert any("J" in r for r in cf_ranges)
    assert any("K" in r for r in cf_ranges)
    assert any("N" in r for r in cf_ranges)
    assert not any("AM4" in r or "AR4" in r for r in cf_ranges)
    h_cf = next(
        rules[0]
        for rng, rules in ws.conditional_formatting._cf_rules.items()
        if str(rng).startswith("<ConditionalFormatting H")
    )
    assert h_cf.type == "expression"
    assert "H4<3" in h_cf.formula[0]

    wb_values = load_workbook(out, data_only=False)
    raw_ws = wb_values[RAW_SHEET]
    assert raw_ws["A1"].value == "directory_name"
    assert raw_ws["A2"].value == "Example Judge"


def test_write_workbook_omits_rule_error_columns(tmp_path: Path):
    df = pd.DataFrame(
        [
            {
                "directory_name": "Dance Judge",
                "international_judge": False,
                "mbr_number": "654321",
                "us_champs_senior_availability": "Available",
                "appointment_year": 2015,
                "last_champs_in_role": 2024,
                "last_sectionals_in_role": 2021,
                "total_comps_in_role_2yr": 8,
                "activity_competition_count": 3,
                "activity_segment_count": 22,
                "activity_junior_senior_segment_count": 18,
                "pcs_marking_score": 0.8,
                "element_marking_score": 0.7,
                "competition_count": 3,
                "segment_count": 22,
                "junior_senior_segment_count": 18,
                "anomaly_rate_pct": 0.8,
                "champs_competition_count": 1,
                "champs_segment_count": 4,
                "champs_junior_senior_segment_count": 4,
                "champs_anomaly_rate_pct": 0.5,
            }
        ]
    )
    out = tmp_path / "dance.xlsx"
    write_national_sp_judge_analysis_xlsx(
        df, out, include_rule_errors=False, sectionals_activity_min_year=2023
    )

    wb = load_workbook(out, data_only=False)
    ws = wb[ANALYSIS_SHEET]
    assert ws.column_dimensions["L"].hidden is True
    assert ws.column_dimensions["U"].hidden is True
    assert ws.column_dimensions["AD"].hidden is True
    assert ws.column_dimensions["AR"].hidden is True
    assert ws.column_dimensions["AW"].hidden is True
    assert ws.column_dimensions["BB"].hidden is True
    assert ws["AV4"].value.startswith("=IF(AQ4")
    assert "AS4:AU4" in str(ws["AV4"].value)
    assert ws["BF4"].value.count("(") == ws["BF4"].value.count(")")
    assert ws["AP4"].value == '=IF(AND(NOT(B4="X"),OR(G4="",G4<2023)),"Low","")'


def test_pool_judge_stats_sums_across_labels():
    totals = {
        "Old Name": {"scores": 10, "anomalies": 1, "rule_errors": 0, "excess_anomalies": 2},
        "New Name": {"scores": 5, "anomalies": 2, "rule_errors": 1, "excess_anomalies": 1},
    }
    pooled = _pool_judge_stats(totals, ["Old Name", "New Name"])
    assert pooled["scores"] == 15
    assert pooled["excess_anomalies"] == 3


def test_write_workbook_dance_layout_labels(tmp_path: Path):
    from scripts.national_judge_report_thresholds import ICE_DANCE_THRESHOLDS

    df = pd.DataFrame(
        [
            {
                "directory_name": "Dance Judge",
                "total_comps_in_role_2yr": 8,
                "activity_competition_count": 3,
                "activity_segment_count": 22,
                "activity_junior_senior_segment_count": 18,
                "competition_count": 3,
                "segment_count": 22,
                "junior_senior_segment_count": 18,
                "anomaly_rate_pct": 0.8,
                "champs_competition_count": 0,
            }
        ]
    )
    out = tmp_path / "dance_layout.xlsx"
    write_national_sp_judge_analysis_xlsx(
        df,
        out,
        include_rule_errors=False,
        thresholds=ICE_DANCE_THRESHOLDS,
        performance_block_header=(
            "Competition Performance (Past three years, all activity)"
        ),
        activity_column_label="Competition Activity",
        performance_analysis_header="Performance Analysis",
        recent_period_header="Last 3 years (all activity)",
    )
    ws = load_workbook(out)["analysis"]
    assert "all activity" in str(ws["I2"].value)
    assert ws["H2"].value == "Last 3 years (all activity)"
    assert ws["AN3"].value == "Competition Activity"
    assert ws["AR2"].value == "Performance Analysis"
    assert f"<={ICE_DANCE_THRESHOLDS.total_comps_in_role_low}" in str(
        ws["AM4"].value
    )


def test_deviation_benchmark_kwargs_all_competitions_all_levels_for_junior_senior_ranking():
    kw = _deviation_benchmark_kwargs(
        start_season="2324",
        end_season="2526",
        segment_level_preset="junior_senior",
    )
    assert kw["benchmark_competition_scope_key"] == COMPETITION_SCOPE_ALL
    assert kw["benchmark_segment_level_preset"] == ELEMENT_RANKING_LEVEL_FILTER_ALL


def test_deviation_benchmark_kwargs_unrestricted_when_ranking_has_no_level_preset():
    kw = _deviation_benchmark_kwargs(
        start_season="2324",
        end_season="2526",
        segment_level_preset=None,
    )
    assert kw["benchmark_competition_scope_key"] == COMPETITION_SCOPE_ALL
    assert "benchmark_segment_level_preset" not in kw


def test_segment_counts_for_judge_ids_can_filter_competition_ids():
    by_judge = {
        1: {(100, 1, "Senior"), (200, 2, "Junior")},
    }
    all_counts = _segment_counts_for_judge_ids(by_judge, {1})
    filtered = _segment_counts_for_judge_ids(by_judge, {1}, competition_ids={100})
    assert all_counts["competition_count"] == 2
    assert filtered["competition_count"] == 1


def test_segment_counts_for_judge_ids_unions_distinct_competitions():
    """Overlapping judge ids must not double-count competitions or segments."""
    by_judge = {
        1: {(100, 1, "Senior"), (100, 2, "Junior"), (200, 3, "Senior")},
        2: {(100, 1, "Senior"), (300, 4, "Senior")},
        99: {(900, 99, "Senior")},
    }
    counts = _segment_counts_for_judge_ids(by_judge, {1, 2})
    assert counts["competition_count"] == 3
    assert counts["segment_count"] == 4
    assert counts["junior_senior_segment_count"] == 4


def test_segment_counts_for_judge_ids_excludes_small_synchro_fields():
    by_judge = {
        1: {
            (100, 1, "Senior"),
            (100, 2, "Junior"),
            (200, 3, "Senior"),
        },
    }
    team_counts = {1: 8, 2: 1, 3: 4}
    counts = _segment_counts_for_judge_ids(
        by_judge,
        {1},
        segment_team_counts=team_counts,
        junior_senior_min_team_count=3,
    )
    assert counts["segment_count"] == 3
    assert counts["junior_senior_segment_count"] == 2


def test_write_workbook_synchro_layout_labels(tmp_path: Path):
    df = pd.DataFrame(
        [
            {
                "directory_name": "Synchro Judge",
                "total_comps_in_role_2yr": 8,
                "activity_competition_count": 3,
                "activity_segment_count": 22,
                "activity_junior_senior_segment_count": 18,
                "competition_count": 3,
                "segment_count": 22,
                "junior_senior_segment_count": 6,
                "anomaly_rate_pct": 0.8,
                "champs_competition_count": 0,
            }
        ]
    )
    synchro = REPORT_DISCIPLINE_CONFIGS["synchro"]
    out = tmp_path / "synchro_layout.xlsx"
    write_national_sp_judge_analysis_xlsx(
        df,
        out,
        include_rule_errors=False,
        thresholds=synchro.thresholds,
        performance_block_header=synchro.performance_block_header,
        activity_column_label=synchro.activity_column_label,
        performance_analysis_header=synchro.performance_analysis_header,
        recent_period_header=synchro.recent_period_header,
        junior_senior_segment_count_header=synchro.junior_senior_segment_count_header,
        junior_senior_activity_label=synchro.junior_senior_activity_label,
        sectionals_block_header=f"{synchro.sectionals_label} (since 2018 for GOEs and 2022 for PCS)",
        sectionals_performance_header=f"{synchro.sectionals_label} Performance Analysis",
    )
    ws = load_workbook(out)["analysis"]
    assert ws["K3"].value == SYNCHRO_JUNIOR_SENIOR_SEGMENT_COUNT_HEADER
    assert ws["T3"].value == SYNCHRO_JUNIOR_SENIOR_SEGMENT_COUNT_HEADER
    assert ws["AC3"].value == SYNCHRO_JUNIOR_SENIOR_SEGMENT_COUNT_HEADER
    assert ws["AO3"].value == SYNCHRO_JUNIOR_SENIOR_ACTIVITY_LABEL
    assert "Synchronized Sectionals" in str(ws["R2"].value)
    assert load_workbook(out)[LOOKUP_SHEET]["A4"].value == synchro.thresholds.element_marking_score_fair


def test_official_candidate_judge_ids_uses_link_not_copanel():
    """Only linked/name-matched judge ids — not every judge on shared segments."""
    linked = {42: {7, 8}}
    groups = {42: {7}}
    match_keys = {"sharon rogers": {7, 99, 100}}
    candidates = _official_candidate_judge_ids(
        official_id=42,
        full_name="Sharon Rogers",
        linked_jids_by_official=linked,
        match_key_to_judge_ids=match_keys,
        group_judge_ids_by_official=groups,
    )
    assert candidates == {7, 8, 99, 100}
    unrelated = _official_candidate_judge_ids(
        official_id=99,
        full_name="Nobody Here",
        linked_jids_by_official={},
        match_key_to_judge_ids=match_keys,
        group_judge_ids_by_official={},
    )
    assert unrelated == set()
