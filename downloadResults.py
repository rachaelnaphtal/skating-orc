import asyncio
from io import BytesIO
from pyppeteer import launch
import gcp_interactions_helper
import judgingParsing
from judgingParsing import autofit_worksheet
from sharedJudgingAnalysis import format_out_of_range_sheets
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import openpyxl
from urllib.parse import urljoin
import time
import pdfkit
from database import test_connection, get_db_session
from database_loader import DatabaseLoader
from sqlalchemy.orm import Session
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
    Color,
)
from google.cloud import storage
from gcp_interactions_helper import write_file_to_gcp
from gcp_interactions_helper import save_gcp_workbook


def convert_url_to_pdf(url, pdf_path):
    try:
        pdfkit.from_url(url, pdf_path)
        # print(f"PDF generated and saved at {pdf_path}")
    except Exception as e:
        print(f"PDF generation failed: {e}")

def download_pdf(url, pdf_path, use_gcp=False):
    r = requests.get(url, timeout=30)
    r.raise_for_status()

    if use_gcp:
        write_file_to_gcp(r.content, pdf_path)
    else:
        with open(pdf_path, "wb") as f:
            f.write(r.content)

async def generate_pdf(url, pdf_path, use_gcp=False):
    browser = await launch(
        {
            "autoClose": False,
            "handleSIGINT": False,
            "handleSIGTERM": False,
            "handleSIGHUP": False,
        }
    )
    page = await browser.newPage()
    await page.goto(url)
    pdf_data = await page.pdf({"format": "A4"})
    if use_gcp:
        write_file_to_gcp(pdf_data, pdf_path)
    else:
        with open(pdf_path, "wb") as f:
            f.write(pdf_data)
    await browser.close()


def processEvent(
    url,
    eventName,
    judges,
    workbook,
    pdf_number,
    event_regex,
    pdf_folder,
    excel_path,
    only_rule_errors=False,
    use_gcp=False,
    create_thrown_out_analysis=False,
    use_html=True,
    judge_filter="",
    isFSM=False
):
    pdf_path = f"{pdf_folder}{eventName}.pdf"
    if isFSM:
        download_pdf(url, pdf_path, use_gcp=use_gcp)
        return judgingParsing.extract_judge_scores(
            workbook=workbook,
            pdf_path=pdf_path,
            base_excel_path=excel_path,
            judges=judges,
            pdf_number=pdf_number,
            event_regex=event_regex,
            only_rule_errors=only_rule_errors,
            url=url,
            use_gcp=use_gcp,
            create_thrown_out_analysis=create_thrown_out_analysis,
            judge_filter=judge_filter,
            use_html=use_html,
            isFSM=True
        )
    if use_html:
        return judgingParsing.extract_judge_scores(
            workbook=workbook,
            pdf_path=pdf_path,
            base_excel_path=excel_path,
            judges=judges,
            pdf_number=pdf_number,
            event_regex=event_regex,
            only_rule_errors=only_rule_errors,
            url=url,
            use_gcp=use_gcp,
            create_thrown_out_analysis=create_thrown_out_analysis,
            judge_filter=judge_filter,
            use_html=use_html
        )
    asyncio.run(generate_pdf(url, pdf_path, use_gcp=use_gcp))
    # convert_url_to_pdf(url, pdf_path)
    return judgingParsing.extract_judge_scores(
        workbook,
        pdf_path,
        excel_path,
        judges,
        pdf_number,
        event_regex,
        only_rule_errors,
        use_gcp=use_gcp,
        create_thrown_out_analysis=create_thrown_out_analysis,
        judge_filter=judge_filter,
        use_html=False
    )


def get_page_contents(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36"
    }

    page = requests.get(url, headers=headers)

    if page.status_code == 200:
        return page.text

    return None

#### FSM parsing ####
def _find_judges_scores_pdf_anchor(td):
    """Swiss Timing FSM: anchor text is often split across nodes; avoid ``string=`` matcher."""
    if td is None:
        return None
    for a in td.find_all("a", href=True):
        if "Judges Scores" in a.get_text():
            return a
    return None


def _role_is_panel_judge(role_label: str) -> bool:
    r = (role_label or "").strip().lower()
    return bool(r) and r.startswith("judge")


def get_fsm_judges_and_results_links(page_contents, base_url):
    soup = BeautifulSoup(page_contents, "html.parser")

    panels = []
    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) != 5:
            continue

        scores_link = _find_judges_scores_pdf_anchor(tds[4])
        panel_link = tds[2].find("a", href=True)

        if not scores_link or not panel_link:
            continue

        panel_url = urljoin(base_url, panel_link["href"])
        panel_html = get_page_contents(panel_url)
        segment_official_rows = (
            parse_ijs_segment_officials(panel_html) if panel_html else []
        )
        judges = [
            r["name"] for r in segment_official_rows if _role_is_panel_judge(r["role"])
        ]

        panels.append({
            "judges": judges,
            "scores_url": urljoin(base_url, scores_link["href"]),
            "segment_official_rows": segment_official_rows,
            "panel_url": panel_url,
        })

    return panels


def parse_judges_from_panel(panel_url):
    """Judge names only (for PDF deviation logic); uses the same parsing as segment officials."""
    panel_html = get_page_contents(panel_url)
    if not panel_html:
        return []
    rows = parse_ijs_segment_officials(panel_html)
    return [r["name"] for r in rows if _role_is_panel_judge(r["role"])]

#### IJS companion parsing
def get_urls_and_names(page_contents):
    soup = BeautifulSoup(page_contents, "html.parser")

    links = soup.find_all("a", href=True, string="Final")
    names = soup.find_all("td", class_="event tRow bRow")
    return list(dict.fromkeys(links)), names


def iter_ijs_index_final_href_and_cover_event(page_contents: str):
    """
    For each ``Final`` link on an IJS index.asp, yield ``(href, cover_event_label)``.
    ``cover_event_label`` is the text of the row's ``<td class="... event ...">`` (short program / free skate).
    Hrefs are yielded at most once each (document order).
    """
    soup = BeautifulSoup(page_contents, "html.parser")
    seen_href: set[str] = set()
    for tr in soup.find_all("tr"):
        final_a = tr.find("a", href=True, string="Final")
        if not final_a:
            continue
        href = (final_a.get("href") or "").strip()
        if not href or href in seen_href:
            continue
        seen_href.add(href)
        event_td = tr.find("td", class_=lambda c: bool(c) and "event" in c)
        cover = event_td.get_text(strip=True) if event_td else ""
        yield href, cover


def iter_fsm_leaderboard_panel_href_and_cover_event(page_contents: str):
    """
    USFS / Swiss Timing leaderboard ``index.htm``: rows with ``Panel of Judges`` and
    ``Judges Scores (pdf)``. Yields ``(panel_of_href, cover_label)`` where
    ``cover_label`` is ``"{category} - {segment}"`` (e.g. Championship Men - Short Program)
    for alignment with ``ijs_event_label_to_db_segment_name``.
    """
    soup = BeautifulSoup(page_contents, "html.parser")
    current_category = ""
    seen_href: set[str] = set()
    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 3:
            ent_a = tds[2].find("a", href=True)
            if ent_a:
                ht = (ent_a.get("href") or "").upper()
                txt = ent_a.get_text(strip=True).lower()
                if ("CAT" in ht and "EN.HTM" in ht) or "entries" in txt:
                    cat = tds[0].get_text(strip=True)
                    if cat:
                        current_category = cat
        if len(tds) != 5:
            continue
        scores_link = _find_judges_scores_pdf_anchor(tds[4])
        panel_a = tds[2].find("a", href=True)
        if not scores_link or not panel_a:
            continue
        href = (panel_a.get("href") or "").strip()
        if not href or href in seen_href:
            continue
        ph = href.upper()
        if not ph.endswith("OF.HTM") and not ph.endswith("OF.HTML"):
            continue
        seen_href.add(href)
        segment_name = tds[1].get_text(strip=True)
        if current_category and segment_name:
            cover = f"{current_category} - {segment_name}"
        else:
            cover = segment_name or current_category
        yield href, cover


def _parse_fsm_function_name_officials(soup) -> list[dict]:
    """Swiss Timing / USFS ``Panel of Judges`` pages (Function / Name table)."""
    out: list[dict] = []
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        header_cells = rows[0].find_all(["th", "td"])
        if len(header_cells) < 2:
            continue
        h0 = header_cells[0].get_text(strip=True).lower()
        h1 = header_cells[1].get_text(strip=True).lower()
        if h0 != "function" or "name" not in h1:
            continue
        for tr in rows[1:]:
            if tr.find("th"):
                continue
            tds = tr.find_all("td")
            if len(tds) < 2:
                continue
            role = tds[0].get_text(strip=True)
            name_cell = tds[1].get_text(strip=True)
            if not role or not name_cell:
                continue
            name_only = name_cell.split(",")[0].strip()
            name_only = name_only.replace("Ms. ", "").replace("Mr. ", "")
            out.append({"role": role, "name": name_only})
        if out:
            return out
    return []


def _disambiguate_official_roles(rows: list[dict]) -> list[dict]:
    """FSM panels may list two ``Technical Specialist`` rows; DB requires unique ``role`` per segment."""
    seen: dict[str, int] = {}
    out: list[dict] = []
    for r in rows:
        base = (r.get("role") or "").strip()
        name = r.get("name")
        n = seen.get(base, 0) + 1
        seen[base] = n
        role = f"{base} ({n})" if n > 1 else base
        out.append({"role": role, "name": name})
    return out


def parse_ijs_segment_officials(page_contents: str) -> list[dict]:
    """
    Parse officials on a segment page: classic IJS ``table.officials``, or FSM
    ``Function`` / ``Name`` panel table (e.g. SEG001OF.htm on USFS leaderboard).
    Returns dicts: ``role`` (e.g. Judge No.1, Referee), ``name`` (short name before comma).
    """
    if not page_contents:
        return []
    soup = BeautifulSoup(page_contents, "html.parser")
    table = soup.find("table", class_=lambda c: c and "officials" in str(c).split())
    if table is not None:
        body = table.find("tbody") or table
        out = []
        for tr in body.find_all("tr"):
            if tr.find("th"):
                continue
            tds = tr.find_all("td")
            if len(tds) < 2:
                continue
            role = tds[0].get_text(strip=True)
            name_cell = tds[1].get_text(strip=True)
            if not role or not name_cell:
                continue
            name_only = name_cell.split(",")[0].strip()
            name_only = name_only.replace("Ms. ", "").replace("Mr. ", "")
            out.append({"role": role, "name": name_only})
        if out:
            return _disambiguate_official_roles(out)
    return _disambiguate_official_roles(_parse_fsm_function_name_officials(soup))


def _classic_segment_official_rows(base_url: str, segment_href: str) -> list[dict]:
    """
    Officials for classic ``index.asp`` flow: try the Final results page, then a Swiss Timing
    style ``*OF.htm`` panel page (e.g. ``SEG012OF.htm`` from ``SEG012.htm``).
    """
    base = (base_url or "").rstrip("/")
    href = (segment_href or "").strip()
    if not href or not base:
        return []
    urls: list[str] = [f"{base}/{href}"]
    m = re.match(r"(?is)^(.+?)(\.(htm|html))$", href)
    if m and not m.group(1).upper().endswith("OF"):
        urls.append(f"{base}/{m.group(1)}OF{m.group(2)}")
    for u in urls:
        html = get_page_contents(u)
        rows = parse_ijs_segment_officials(html or "")
        if rows:
            # if u != urls[0]:
            #     print(f"INFO: Panel officials loaded from {u!r} (OF fallback)")
            return rows
    return []


def findJudgesNames(soup):
    alltd = soup.find_all("td")
    judges = []
    nextJudge = False
    judgeNumber = 1
    for td in alltd:
        if td.text.count("Judge ") > 0:
            nextJudge = True
            judgeNumber = int(td.text.replace("Judge ", ""))
            while len(judges) < judgeNumber:
                judges.append("")
        elif nextJudge:
            judges[judgeNumber-1] = td.text.split(",")[0]
            nextJudge = False
    return judges


# event_name -> total_errors_per_judge, allowed_errors
def make_competition_summary_page(
    workbook,
    report_name,
    event_details_dict,
    judge_errors,
    only_rule_errors=False,
    use_gcp=False,
):
    sheet = workbook.create_sheet("Summary", 0)

    # Styles
    bold = Font(bold=True)
    gray = PatternFill("solid", fgColor="C0C0C0")
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrap_text = Alignment(wrap_text=True)
    vertical_text = Alignment(textRotation=90)

    # sheet.freeze_panes("D1")
    sheet.cell(1, 1, value=report_name)
    sheet.cell(1, 1).font = bold
    sheet.cell(2, 1, value="Official's Review Summary")
    sheet.cell(2, 1).font = bold
    sheet.cell(6, 1, value="EVENT")
    sheet.cell(6, 1).alignment = wrap_text
    sheet.cell(6, 1).border = thin_border
    sheet.cell(6, 2, value="STARTS")
    sheet.cell(6, 2).alignment = wrap_text
    sheet.cell(6, 2).border = thin_border
    sheet.cell(6, 3, value="ALLOWED ERRORS")
    sheet.cell(6, 3).alignment = wrap_text
    sheet.cell(6, 3).border = thin_border
    sheet.column_dimensions["A"].width = 35

    current_row = 9
    events_in_order = sorted(event_details_dict.items())
    for event in events_in_order:
        sheet.cell(current_row, 1, value=event[0])
        sheet.cell(current_row, 2, value=event[1]["Num Starts"])
        sheet.cell(current_row, 3, value=event[1]["Allowed Errors"])
        current_row += 1

    current_row += 1
    sheet.cell(current_row, 1, value="TOTALS")
    sheet.cell(current_row, 1).font = bold

    totals_row = current_row
    for i in range(7, totals_row):
        sheet.cell(i, 1).border = Border(right=thin)
        sheet.cell(i, 2).border = Border(right=thin)
        sheet.cell(i, 3).border = Border(right=thin)
    sheet.cell(totals_row, 1).border = thin_border
    sheet.cell(totals_row, 2).border = thin_border
    sheet.cell(totals_row, 3).border = thin_border

    current_col = 4
    for judge in dict(sorted(judge_errors.items())):
        sheet.cell(6, current_col).value = judge
        sheet.cell(6, current_col + 1).alignment = Alignment(
            wrap_text=True, horizontal="center"
        )
        sheet.merge_cells(
            start_row=6, start_column=current_col, end_row=6, end_column=current_col + 3
        )
        sheet.cell(7, current_col, value="Number Anomalies")
        sheet.cell(7, current_col + 1, value="OAC Recognized Errors")
        sheet.cell(7, current_col + 2, value="Errors in Excess Pre-Review")
        sheet.cell(7, current_col + 3, value="Errors in Excess After Review")
        for i in range(4):
            sheet.cell(7, current_col + i).alignment = vertical_text
        current_row = 9
        for event in events_in_order:
            if event[0] in judge_errors[judge]:
                judge_number = judge_errors[judge][event[0]]["Judge Number"]
                sheet_name = event[1]["Sheet Name"]
                row = judge_number + event[1]["Summary Row Start"] - 1
                sheet.cell(
                    current_row,
                    current_col,
                    value=judge_errors[judge][event[0]]["Errors"],
                )
                sheet.cell(
                    current_row, current_col + 1
                ).value = f"='{sheet_name}'!C{row}"
                # print (f"row:{current_row}, col:{current_col+1} value:{sheet.cell(current_row, current_col+1).value}")
                sheet.cell(
                    current_row,
                    current_col + 2,
                    value=judge_errors[judge][event[0]]["In Excess"],
                )
                sheet.cell(current_row, current_col + 2).font = Font(
                    b=True, color="FF0000"
                )
                sheet.cell(
                    current_row, current_col + 3
                ).value = f"=MAX({get_column_letter(current_col + 1)}{current_row}-C{current_row}, 0)"
                sheet.cell(current_row, current_col + 3).font = Font(
                    b=True, color="FF0000"
                )
                for i in range(4):
                    sheet.cell(current_row, current_col + i).fill = PatternFill(
                        "solid", fgColor="CCE5FF"
                    )
            current_row += 1

        for i in range(4, current_col + 4):
            column_letter = get_column_letter(i)
            sheet.cell(
                totals_row, i
            ).value = f"=SUM({column_letter}9:{column_letter}{totals_row - 1})"
        sheet.cell(totals_row, current_col + 3).fill = PatternFill(
            "solid", fgColor="66B2FF"
        )

        # Add borders
        sheet.cell(6, current_col).border = thin_border
        sheet.cell(7, current_col).border = thin_border
        for i in range(8, totals_row):
            sheet.cell(i, current_col).border = Border(left=thin)
            sheet.cell(i, current_col + 3).border = Border(right=thin)
        for i in range(4):
            sheet.cell(6, current_col + i).border = thin_border
            sheet.cell(7, current_col + i).border = thin_border
            sheet.cell(totals_row, current_col + i).border = thin_border

        current_col += 4


def make_old_summary_sheet(workbook, df_dict, judge_errors, event_regex):
    # Add summary sheet
    sheet = workbook.create_sheet("Summary", 0)

    sheet.cell(1, 1, value="Summary")
    current_col = 1
    summary_row = 6 + max([len(judge_errors[judge]) for judge in judge_errors])

    # Add summary row for all anomalies
    sheet.cell(2, current_col, value="Judge Name")
    sheet.cell(2, current_col + 1, value="# Anomalies")
    sheet.cell(2, current_col + 2, value="# ORC Errors")
    sheet.cell(2, current_col + 3, value="# In Excess Anomalies")
    sheet.cell(2, current_col + 4, value="# In Excess After ORC")
    sheet.cell(2, current_col + 5, value="# Events")
    current_row = 3

    for judge, value in sorted(
        df_dict.items(), key=lambda kv: kv[1]["Errors"].sum(), reverse=True
    ):
        sheet.cell(current_row, current_col, value=judge)
        sheet.cell(
            current_row, current_col + 1, value=int(df_dict[judge]["Errors"].sum())
        )
        sheet.cell(
            current_row, current_col + 3, value=int(df_dict[judge]["In Excess"].sum())
        )
        sheet.cell(current_row, current_col + 5,
                   value=len(judge_errors[judge]))
        current_row += 1
    current_col += 7
    for judge in dict(sorted(judge_errors.items())):
        current_row = 2
        sheet.cell(current_row, current_col, value=judge)
        current_row += 1
        sheet.cell(current_row, current_col, value="Event")
        sheet.cell(current_row, current_col + 1, value="Anomalies")
        sheet.cell(current_row, current_col + 2, value="ORC recognized")
        sheet.cell(current_row, current_col + 3, value="Allowed")
        sheet.cell(current_row, current_col + 4,
                   value="In Excess (Pre Review)")
        sheet.cell(current_row, current_col + 5,
                   value="In Excess (Post Review)")
        current_row += 1

        for event in judge_errors[judge]:
            if (
                event == "Total Errors"
                or event == "Allowed Errors"
                or not re.match(event_regex, event)
            ):
                continue
            sheet.cell(current_row, current_col, event.replace("_", " "))
            num_errors = judge_errors[judge][event]["Errors"]
            sheet.cell(current_row, current_col + 1, value=num_errors)

            num_allowed = int(judge_errors[judge][event]["Allowed Errors"])
            sheet.cell(current_row, current_col + 3, value=num_allowed)

            in_excess = int(judge_errors[judge][event]["In Excess"])
            sheet.cell(current_row, current_col + 4, value=in_excess)
            current_row += 1

        current_row += 1

        sheet.cell(summary_row, current_col, value="Total")
        sheet.cell(
            summary_row, current_col + 1, value=int(df_dict[judge]["Errors"].sum())
        )
        sheet.cell(
            summary_row, current_col + 4, value=int(df_dict[judge]["In Excess"].sum())
        )
        current_col += 5
    judgingParsing.autofit_worksheet(sheet)


def findResultsDetailUrlAndJudgesNames(base_url, results_page_link):
    url = f"{base_url}/{results_page_link}"
    page_contents = get_page_contents(url)
    if not page_contents:
        print(f"WARNING: Empty or failed HTML fetch for Final page {url!r}")
        return ("", [], "")
    soup = BeautifulSoup(page_contents, "html.parser")
    link = soup.find("a", href=True, string="Judge detail scores")
    judgesNames = findJudgesNames(soup)
    event_name = soup.find_all('h1')[0].get_text()
    details_link = ""
    if link is not None:
        details_link = link["href"]
    return (details_link, judgesNames, event_name)


def loadCompetitionInfo(base_url):
    if base_url.endswith(".htm"):
        return loadCompetitionInfoFSM(base_url)
    
    page_contents = get_page_contents(base_url)
    soup = BeautifulSoup(page_contents, "html.parser")
    all_h3_tags = soup.find_all('h3')
    date = all_h3_tags[0].get_text().split(" ")
    start_date = date[0]
    end_date = date[2]
    location = all_h3_tags[2].get_text()
    return (start_date, end_date, location)

def loadCompetitionInfoFSM(base_url):
    page_contents = get_page_contents(base_url)
    soup = BeautifulSoup(page_contents, "html.parser")

    start_date, end_date = soup.find_all("tr", class_="caption3")[0].find("td").text.replace(" ", "").split("-")
    location = soup.find_all("td", class_="caption3")[0].text
    return (start_date, end_date, location)


def loadInfoForExistingCompetitions():
    session = get_db_session()
    database_obj = DatabaseLoader(session)
    urls = database_obj.getCompetitionUrlsWithNoLocation()
    for url in urls:
        (start_date, end_date, location) = loadCompetitionInfo(
            f"{url}/index.asp")
        database_obj.updateCompetition(
            url, location=location, start_date=start_date, end_date=end_date)



def scrape(
    base_url,
    report_name,
    excel_folder="",
    pdf_folder="/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/Easterns/Results/",
    event_regex="",
    only_rule_errors=False,
    use_gcp=False,
    add_additional_analysis=False,
    write_to_database=False,
    year="2526",
    judge_filter="",
    specific_exclude="",
    use_html=True,
    isFSM=False
):
    """
    When ``write_to_database`` is true, each ``public.segment`` row is named from the score
    parse only: FSM uses ``process_fsm_scores`` / ``parse_scores`` (PDF header); classic IJS
    uses ``process_scores_html`` (detail page ``catseg``). Panel officials are written after a
    successful score parse. If a segment is skipped (e.g. ``event_regex``) but the parsed
    name already matches an existing segment with no ``segment_official`` rows, panel data is
    loaded when available. Index-only labels from ``segment_officials_backfill`` are not used.
    """
    url = f"{base_url}/index.asp"
    if isFSM:
        url = f"{base_url}/index.htm"
    page_contents = get_page_contents(url)
    print(url)
    workbook = openpyxl.Workbook()
    agg_all_element_df = None
    agg_all_pcs_df = None
    session = get_db_session()
    database_obj = DatabaseLoader(session)

    competition_id=0
    proccessed_segments=[]
    if write_to_database:
        competition_id = database_obj.insert_competition(
            report_name.replace("_", " "), base_url, year)
        proccessed_segments = database_obj.getSegmentNamesForCompetition(
            base_url)
        (start_date, end_date, location) = loadCompetitionInfo(
            url)
        database_obj.updateCompetition(
            base_url, location=location, start_date=start_date, end_date=end_date)

    if page_contents:
        judge_errors = {}
        event_details = {}
        detailed_rule_errors = []
        if isFSM:
            judges_and_results_links = get_fsm_judges_and_results_links(page_contents, f"{base_url}/")
            i=0
            for event_info_dict in judges_and_results_links:
                judges = event_info_dict["judges"]
                scores_url = event_info_dict["scores_url"]
                (
                    event_name,
                    total_errors,
                    num_starts,
                    allowed_errors,
                    rule_errors,
                    all_element_dict,
                    all_pcs_dict,
                ) = processEvent(
                    scores_url,
                    i,
                    judges,
                    workbook,
                    i,
                    event_regex,
                    pdf_folder,
                    excel_folder,
                    only_rule_errors=only_rule_errors,
                    use_gcp=use_gcp,
                    create_thrown_out_analysis=add_additional_analysis or write_to_database,
                    judge_filter=judge_filter,
                    use_html=use_html,
                    isFSM=True
                )
                segment_official_rows = None
                if write_to_database:
                    segment_official_rows = (
                        event_info_dict.get("segment_official_rows") or None
                    )
                agg_all_element_df, agg_all_pcs_df = handleEventResults(report_name, write_to_database, 
                                                                        judge_filter, agg_all_element_df, agg_all_pcs_df, 
                                                                        database_obj, competition_id, proccessed_segments, 
                                                                        judge_errors, event_details, detailed_rule_errors, 
                                                                        i, judges, event_name, total_errors, num_starts, 
                                                                        allowed_errors, rule_errors, all_element_dict, all_pcs_dict,
                                                                        segment_official_rows=segment_official_rows)
                i=i+1
        else:
            links, names = get_urls_and_names(page_contents)
            for i in range(len(links)):
                segment_href = links[i]["href"]
                (resultsLink, judgesNames, h1_event_label) = findResultsDetailUrlAndJudgesNames(
                    base_url, segment_href
                )
                if not resultsLink:
                    print(
                        f"WARNING: No judge detail scores link on Final page {segment_href!r}; skipping"
                    )
                    continue
                if specific_exclude and (h1_event_label == specific_exclude or re.match(specific_exclude, h1_event_label)):
                    continue

                (
                    event_name,
                    total_errors,
                    num_starts,
                    allowed_errors,
                    rule_errors,
                    all_element_dict,
                    all_pcs_dict,
                ) = processEvent(
                    f"{base_url}/{resultsLink}",
                    i,
                    judgesNames,
                    workbook,
                    i,
                    event_regex,
                    pdf_folder,
                    excel_folder,
                    only_rule_errors=only_rule_errors,
                    use_gcp=use_gcp,
                    create_thrown_out_analysis=add_additional_analysis or write_to_database,
                    judge_filter=judge_filter,
                    use_html=use_html
                )
                segment_official_rows = None
                segment_db_key = None
                if write_to_database:
                    segment_official_rows = _classic_segment_official_rows(
                        base_url, segment_href
                    )
                    if not segment_official_rows:
                        print(
                            f"WARNING: No panel officials parsed for Final {segment_href!r}"
                        )
                    segment_db_key = (
                        (event_name or "").strip()
                        or judgingParsing.ijs_event_label_to_db_segment_name(
                            h1_event_label or ""
                        )
                    )
                agg_all_element_df, agg_all_pcs_df = handleEventResults(report_name, write_to_database,
                                                                        judge_filter, agg_all_element_df, agg_all_pcs_df,
                                                                        database_obj, competition_id, proccessed_segments,
                                                                        judge_errors, event_details, detailed_rule_errors,
                                                                        i, judgesNames, event_name, total_errors, num_starts,
                                                                        allowed_errors, rule_errors, all_element_dict, all_pcs_dict,
                                                                        segment_official_rows=segment_official_rows,
                                                                        segment_db_key=segment_db_key)
        # Sort sheets
        del workbook["Sheet"]
        workbook._sheets.sort(key=lambda ws: ws.title)

        df_dict = {}
        for judge in judge_errors:
            df_dict[judge] = pd.DataFrame.from_dict(
                judge_errors[judge], orient="index")

        errors_dict_to_return = pd.DataFrame.from_dict(detailed_rule_errors)

        make_competition_summary_page(
            workbook, report_name, event_details, judge_errors
        )

    else:
        print("Failed to get page contents.")

    excel_path = f"{excel_folder}{report_name}.xlsx"
    if use_gcp:
        save_gcp_workbook(workbook, excel_path)
    else:
        workbook.save(excel_path)

    if add_additional_analysis:
        create_additional_analysis_sheet(
            agg_all_element_df, agg_all_pcs_df, excel_folder, report_name, use_gcp=use_gcp
        )
    print("Finished " + report_name + datetime.now().strftime("%H:%M:%S"))
    return df_dict, errors_dict_to_return

def handleEventResults(report_name, write_to_database, judge_filter, agg_all_element_df, agg_all_pcs_df, database_obj, competition_id, proccessed_segments, judge_errors, event_details, detailed_rule_errors, event_number, judgesNames, event_name, total_errors, num_starts, allowed_errors, rule_errors, all_element_dict, all_pcs_dict, segment_official_rows=None, segment_db_key=None):
    row_segment_key = ((segment_db_key or event_name) or "").strip()
    if total_errors == None:
        if (
            write_to_database
            and competition_id
            and segment_official_rows
            and row_segment_key
        ):
            segment_id = database_obj.get_segment_id(
                row_segment_key, competition_id
            )
            if segment_id is not None:
                database_obj.replace_segment_officials(
                    segment_id, segment_official_rows
                )
                print(
                    f" Officials only (skipped segment) {row_segment_key} "
                    f"{datetime.now().strftime('%H:%M:%S')}"
                )
            else:
                print(
                    f"WARNING: Skipped scores for {row_segment_key!r}; no segment row in DB "
                    f"— officials not written"
                )
        return agg_all_element_df, agg_all_pcs_df
    # ``event_name`` must match ``extract_judge_scores`` (parse_scores / process_scores_html).
    start_of_summary_rows = sum(total_errors) + 11
    event_details[event_name] = {
                    "Num Starts": num_starts,
                    "Allowed Errors": allowed_errors,
                    "Sheet Name": judgingParsing.get_sheet_name(event_name, event_number),
                    "Summary Row Start": start_of_summary_rows,
                }
    for event_number in range(len(judgesNames)):
        judge = judgesNames[event_number]
        if len(judge_filter) > 0 and judge_filter != judge:
            continue
        if judge not in judge_errors:
            judge_errors[judge] = {}
        judge_errors[judge][event_name] = {
                        "Errors": total_errors[event_number],
                        "Allowed Errors": allowed_errors,
                        "In Excess": max(total_errors[event_number] - allowed_errors, 0),
                        "Judge Number": event_number + 1,
                    }

    for rule_error in rule_errors:
        rule_error["Competition"] = report_name
        rule_error["Event"] = event_name
        detailed_rule_errors.append(rule_error)

                # Processing of additional info
    all_pcs_df = pd.DataFrame.from_dict(all_pcs_dict)
    all_element_df = pd.DataFrame.from_dict(all_element_dict)
    if agg_all_pcs_df is None:
        agg_all_pcs_df = all_pcs_df
    else:
        agg_all_pcs_df = pd.concat([agg_all_pcs_df, all_pcs_df])

    if agg_all_element_df is None:
        agg_all_element_df = all_element_df
    else:
        agg_all_element_df = pd.concat(
                        [agg_all_element_df, all_element_df])

                # write to database
    if write_to_database:
        segment_id = None
        if event_name not in proccessed_segments:
            print(
                        f"Writing segment {event_name} {datetime.now().strftime("%H:%M:%S")}")
            segment_id = database_obj.insert_segment(
                        event_name, competition_id)
            database_obj.insert_element_scores(
                        judgesNames, all_element_dict, segment_id, rule_errors)
            database_obj.insert_pcs_scores(
                        judgesNames, all_pcs_dict, segment_id)
        else:
            print(
                f"Skipping scores for segment {event_name} "
                f"(already present) {datetime.now().strftime("%H:%M:%S")}"
            )
            # ``insert_segment`` returns existing id when the row is already there (same as
            # insert path); avoids ``get_segment_id`` None if the session/DB view differs.
            segment_id = database_obj.insert_segment(event_name, competition_id)
        if segment_id is not None and segment_official_rows:
            database_obj.replace_segment_officials(segment_id, segment_official_rows)
            print(
                f"INFO: segment_official {len(segment_official_rows)} row(s) → "
                f"segment id {segment_id} ({event_name!r})"
            )
    return agg_all_element_df,agg_all_pcs_df


def create_summary_element_df(df, grouping_name):
    # Filter only thrown-out elements
    df_thrown_out = df[df["Thrown out"] == True]

    # Group by Judge Name and Element Type
    summary = (
        df_thrown_out.groupby(["Judge Name", grouping_name])
        .agg(
            Num_Low=("High", lambda x: (x == False).sum()),
            Num_High=("High", lambda x: (x == True).sum()),
        )
        .reset_index()
    )

    # Get the total number of judged elements per Judge Name & Element Type
    total_judged = (
        df.groupby(["Judge Name", grouping_name])
        .size()
        .reset_index(name="Total # Judged")
    )

    # Merge with thrown-out data
    summary = summary.merge(
        total_judged, on=["Judge Name", grouping_name], how="left")

    # Calculate percentages
    summary["% Low"] = summary["Num_Low"] / summary["Total # Judged"]
    summary["% High"] = summary["Num_High"] / summary["Total # Judged"]
    summary["% Out"] = (summary["Num_Low"] + summary["Num_High"]) / summary[
        "Total # Judged"
    ]

    # Rename columns to match your desired format
    summary.rename(columns={"Num_Low": "# Low",
                   "Num_High": "# High"}, inplace=True)
    return summary


def make_analysis_cover_sheet(workbook):
    sheet = workbook.create_sheet("Overview")

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 150
    sheet.row_dimensions[3].height = 40
    sheet.row_dimensions[15].height = 40

    bold = Font(bold=True)

    sheet.cell(1, 1, value="Overview").font = Font(bold=True, size=18)
    sheet.merge_cells("A3:B4")
    sheet.cell(
        3,
        1,
        value="This workbook contains additional analysis of the judging data, specifically related to deviations. \n The first two sheets show the percentage of GOEs or PCS of each type that are extremes related to the judging panel. The first six columns show the absolute numbers and the final three show the percentages of the total.",
    )
    sheet.cell(3, 1).alignment = Alignment(wrap_text=True)

    sheet.cell(7, 1, value="Definitions:").font = bold
    sheet.cell(8, 1, value="Thrown out:").font = bold
    sheet.cell(9, 1, value="Low vs High:").font = bold

    sheet.cell(
        8,
        2,
        value="The number of scores that are the extremes of the panel. If at least three judges give the same score, it does not count as thrown out.",
    )
    sheet.cell(
        9,
        2,
        value="Low refers to the judge being lower than the average of the panel. The total is the sum of low and high.",
    )
    sheet.cell(8, 2).alignment = Alignment(wrap_text=True)
    sheet.cell(9, 2).alignment = Alignment(wrap_text=True)

    sheet.merge_cells("A12:B12")
    sheet.cell(
        12,
        1,
        value="The later sheets show the distinct scores given before processing. Feel free to filter to dig into specifics more.",
    )


def create_additional_analysis_sheet(
    all_element_df, all_pcs_df, excel_folder, report_name, use_gcp=False
):
    excel_path = f"{excel_folder}{report_name}_Additional_Analysis.xlsx"
    excel_buffer = BytesIO()
    if not use_gcp:
        excel_buffer = excel_path
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        make_analysis_cover_sheet(writer.book)
        # Analyze elements
        summary_goe_df = create_summary_element_df(
            all_element_df, "Element Type")
        summary_goe_df.to_excel(
            writer, sheet_name="GOE Thrown out", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["GOE Thrown out"])

        summary_pcs_df = create_summary_element_df(all_pcs_df, "Component")
        summary_pcs_df.to_excel(
            writer, sheet_name="PCS Thrown out", float_format="%.2f", index=False
        )
        format_out_of_range_sheets(writer.sheets["PCS Thrown out"])

        all_element_df.to_excel(writer, sheet_name="All Elements", index=False)
        writer.sheets["All Elements"].auto_filter.ref = writer.sheets[
            "All Elements"
        ].dimensions
        all_pcs_df.to_excel(writer, sheet_name="All PCS", index=False)
        writer.sheets["All PCS"].auto_filter.ref = writer.sheets["All PCS"].dimensions

        for sheet in writer.sheets:
            if sheet != "Overview":
                judgingParsing.autofit_worksheet(writer.sheets[sheet])

    if use_gcp:
        gcp_interactions_helper.write_file_to_gcp(
            excel_buffer.getvalue(), excel_path)


def create_season_summary(pdf_folder="", excel_folder="", full_report_name="2425Summary", only_rule_errors=False):
    start = time.time()
    workbook = openpyxl.Workbook()
    events = {
        "Eastern_Synchro_Sectionals": "2025/34239",
        "Midwest_Synchro_Sectionals": "2025/34240",
        "Pacific_Coast_Synchro_Sectionals": "2025/34241",
        "Dallas_Classic": "2024/33436",
        "Cactus_Classic": "2024/34414",
        "Peach_Open": "2024/33518",
        "Glacier_Falls": "2024/33519",
        "Lake_Placid": "2024/33491",
        # "Philadelphia": "2024/33453",
        # "Scott_Hamilton": "2024/33501",
        # "Copper_Cup": "2024/33425",
        # "Cup_of_Colorado": "2024/33507",
        # "Skate_the_Lake": "2024/33520",
        # "MiddleAtlantics": "2024/33515",
        # "SkateSF": "2024/33479",
        # "Potomac": "2024/33523",
        # # "Providence": "",
        # "Chicagoland": "2024/33497",
        # "JohnSmith": "2024/33451",
        # "Pasadena": "2024/33509",
        # "PNIW": "2024/33489",
        # "Challenge_Cup": "2024/34444",
        # "Skate_Cleveland": "2024/33466",
        # "Austin_Autumn_Classic": "2024/33458",
        "BostonNQS": "2024/33526",
        "Pacifics2425": "2024/34291",
        "Midwesterns_DanceFinal2425": "2024/34290",
        "Easterns_PairsFinal2425": "2024/34289",
        "2025USChampionships": "2025/35539",

    }
    summary_dict = {}
    all_rules_errors = []

    event_regex = ""
    if only_rule_errors:
        event_regex = ".*Women|Men|Boys|Girls.*"
    for event_name in events:
        start_event = time.time()
        result, rule_errors = scrape(
            f"https://ijs.usfigureskating.org/leaderboard/results/{events[event_name]}",
            f"{event_name}",
            event_regex,
            only_rule_errors=only_rule_errors,
            write_to_database=True,
            pdf_folder=pdf_folder,
            isFSM=True
        )
        all_rules_errors.append(rule_errors)
        for judge in result:
            result[judge]["Competition"] = event_name
            if judge not in summary_dict:
                summary_dict[judge] = result[judge]
            else:
                summary_dict[judge] = pd.concat(
                    [summary_dict[judge], result[judge]], axis=0
                )
        print(f"{time.time() - start_event} seconds for {event_name}")
    print_summary_workbook(workbook, summary_dict, full_report_name)
    print_rule_error_summary_workbook(
        pd.concat(all_rules_errors, ignore_index=False), full_report_name
    )
    print(f"{time.time() - start} seconds elapsed total")


def print_summary_workbook(workbook, summary_dict, full_report_name):
    sheet = workbook.create_sheet("Summary", 0)
    # Specifying style
    # bold = xlwt.easyxf('font: bold 1')
    sheet.cell(1, 1, value="Summary")
    current_col = 1

    # Add summary sheet for all anomalies
    sheet.cell(2, current_col, value="Judge Name")
    sheet.cell(2, current_col + 1, value="# Anomalies")
    sheet.cell(2, current_col + 2, value="# In Excess")
    sheet.cell(2, current_col + 3, value="# Events")
    sheet.cell(2, current_col + 4, value="In Excess per event")
    current_row = 3

    for judge, value in sorted(
        summary_dict.items(), key=lambda kv: kv[1]["In Excess"].sum(), reverse=True
    ):
        sheet.cell(current_row, current_col, value=judge)
        sheet.cell(current_row, current_col + 1,
                   value=int(value["Errors"].sum()))
        sheet.cell(current_row, current_col + 2,
                   value=int(value["In Excess"].sum()))
        num_events = len(value[value["Errors"] >= 0])
        sheet.cell(current_row, current_col + 3, value=num_events)
        sheet.cell(
            current_row,
            current_col + 4,
            value=float(value["In Excess"].sum()) / float(num_events),
        )
        current_row += 1

    excel_path = f"{excel_folder}{full_report_name}.xls"
    workbook.save(excel_path)
    # Add sheets per judge
    writer = pd.ExcelWriter(
        f"{excel_folder}{full_report_name}_perJudge.xlsx", engine="openpyxl"
    )
    for judge in sorted(summary_dict.keys()):
        print_sheet_per_judge(writer, judge, summary_dict[judge])
    writer.close()


def print_sheet_per_judge(writer, judge_name: str, judge_df):
    judge_df.rename(columns={"Errors": "Anomalies",
                    "Allowed Errors": "Allowed"})
    judge_df.to_excel(writer, sheet_name=judge_name)
    writer.sheets[judge_name].column_dimensions["A"].width = 35
    writer.sheets[judge_name].column_dimensions["B"].width = 12
    writer.sheets[judge_name].column_dimensions["C"].width = 12
    writer.sheets[judge_name].column_dimensions["D"].width = 12
    writer.sheets[judge_name].column_dimensions["E"].width = 30


def print_rule_error_summary_workbook(rule_errors, full_report_name):
    writer = pd.ExcelWriter(
        f"{excel_folder}{full_report_name}_RuleErrors.xlsx", engine="openpyxl"
    )

    grouped_df = rule_errors.groupby("Judge Name").size()
    grouped_df = grouped_df.sort_values(ascending=False)
    grouped_df.to_excel(writer, sheet_name="Summary")
    autofit_worksheet(writer.sheets["Summary"])

    rule_errors.to_excel(writer, sheet_name="All Errors")
    autofit_worksheet(writer.sheets["All Errors"])

    # Add sheets per judge
    for judge in sorted(rule_errors["Judge Name"].unique()):
        judge_df = rule_errors[rule_errors["Judge Name"] == judge]
        judge_df.to_excel(writer, sheet_name=judge)
        autofit_worksheet(writer.sheets[judge])
    writer.close()


# pdf_folder = "/Users/rnaphtal/Documents/JudgingAnalysis/2425/Results/"  # Update with the correct path
# excel_folder = "/Users/rnaphtal/Documents/JudgingAnalysis/2425/"
# base_url = 'https://ijs.usfigureskating.org/leaderboard/results/2024/34290'
# report_name = "Mids2024_ORC_Report"

# #Mids 2425
# pdf_folder = "/Users/rnaphtal/Documents/JudgingAnalysis/2425/Results/"  # Update with the correct path
# excel_folder = "/Users/rnaphtal/Documents/JudgingAnalysis/2425/"
# base_url = 'https://ijs.usfigureskating.org/leaderboard/results/2024/34290'
# report_name = "Mids2425_ORC_Report"

if __name__ == "__main__":
    # #Easterns/ Pairs Final 2425
    # Update with the correct path
    pdf_folder = "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/Easterns/Results/"
    excel_folder = "/Users/rachaelnaphtal/Documents/JudgingAnalysis_Results/Official/"
    base_url = "https://ijs.usfigureskating.org/leaderboard/results/2024/34289"
    # scrape("https://ijs.usfigureskating.org/leaderboard/results/2025/35539", "US_Champs_25")
    # scrape("https://ijs.usfigureskating.org/leaderboard/results/2025/35539", "US_Champs_25_SP", event_regex=".*(Women|Men|Pairs).*")
    # scrape("https://ijs.usfigureskating.org/leaderboard/results/2025/35539", "US_Champs_25_Dance", event_regex=".*(Dance).*")
    # scrape("https://ijs.usfigureskating.org/leaderboard/results/2025/34240", "Midwestern_Synchro_25", event_regex=".*(Novice|Junior|Senior).*")
    # scrape("https://ijs.usfigureskating.org/leaderboard/results/2025/34241", "PacificCoast_Synchro_25", event_regex=".*(Novice|Junior|Senior).*")
    # scrape(
    #     "https://ijs.usfigureskating.org/leaderboard/results/2025/34241",
    #     "Pacific_Coast_Synchro_25",
    #     excel_folder=excel_folder,
    #     pdf_folder=pdf_folder,
    #     # event_regex=".*(Dance).*",
    #     add_additional_analysis=False,
    #     write_to_database=True
    # )
    # scrape("https://ijs.usfigureskating.org/leaderboard/results/2025/34241", "PacificCoast_Synchro_25_all")

    # scrape(base_url, "2024_Pairs_Final_with_errors", ".?(Novice|Junior|Senior).?(Pairs).?")
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2024/34290', "2024_Dance_Final", ".*(Novice|Junior|Senior).?(Dance).*")
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2023/33513', "2023_Boston_NQS", "")
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35701', "PhiladelphiaSummerChallengeNQS2025", "", write_to_database=True, pdf_folder=pdf_folder)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/34237', "2025USSynchroChamps", "", write_to_database=True, pdf_folder=pdf_folder)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35645', "GlacierFallsSummerClassicNQS2025", "", write_to_database=True, pdf_folder=pdf_folder)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35895', "DuPageNQS2025", "", write_to_database=True, pdf_folder=pdf_folder)

    # findResultsDetailUrlAndJudgesNames("https://ijs.usfigureskating.org/leaderboard/results/2025/36275", "CAT023SEG042.html")

    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36275', "DallasClassicNQS2025", "", write_to_database=True, pdf_folder=pdf_folder)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35672', "LakePlacidNQS2025", "", write_to_database=True, pdf_folder=pdf_folder)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35638', "2025SiliconValleyOpen", "", write_to_database=True, pdf_folder=pdf_folder)
    # create_season_summary(pdf_folder=pdf_folder, excel_folder=excel_folder)

    # Issues
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2022/30670', '2022_John_Smith_Memorial_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2223, event_regex=".*(Women|Men|Boys|Girls).*")
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2022/30472', '2022_Cup_of_Colorado_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2223)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2022/30692', '2022_Pasadena_Open_Championships_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2223)
    # # Dance
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2022/30629', '2022_Glacier_Falls_Summer_Classic_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2223)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2022/30690', '2022_Silicon_Valley_Open_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2223)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2022/30693', '2022_Onyx_Challenge_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2223)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2022/30712', '2022_Challenge_Cup_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2223, specific_exclude="Intermediate Men Combined (202 - 209)")

    # 202526 NQS
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36180',
    #        '2025_Skate_Detroit_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35664',
    #        '2025_Atlanta_Open_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35672',
    #        'LP', write_to_database=True, pdf_folder=pdf_folder, year=2526, judge_filter="Jaclyn Helms")
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35103',
    #        '2025_Greater_Chicagoland_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36182',
    #        '2025_Colorado_Springs_Invitational_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35636',
    #        '2025_Carousel_Classic_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35821',
    #        '2025_Maplewood_Classic_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36179',
    #        '2025_Challenge_Cup_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35900',
    #        '2025_Skate_Cleveland_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35890',
    #        '2025_Boston_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/35743',
    #        '2025_Texas_Trophy_NQS', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36292',
    #        '2025_National_Showcase', write_to_database=False, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/34252',
    #        '2026_Eastern_Sectional_Finals', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/34253',
    #        '2026_Midwestern_Sectional_Finals_And_Pairs_Final', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    

    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36375',
    #        '2025_WisconSync', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36347',
    #        '2025_Chuck_Cope', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36286',
    #        '2025_Sunshine_State_Synchro_Classic', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36430',
    #        '2025_Stars_And_Stripes_Synchro_Classic', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36463',
    #        '2025_Capital_Ice_Synchro_Classic', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2026/34247',
    #        '2026_Eastern_Synchronized_Sectional_Championships', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2026/34248',
    #        '2026_Midwestern_Synchronized_Sectional_Championships', write_to_database=True, pdf_folder=pdf_folder, year=2526)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36459',
    #        '2025_Porter_Synchro_Classic', write_to_database=True, pdf_folder=pdf_folder, year=2526, specific_exclude=".*Unified Teams.*")
    
    #FSM testing
    # scrape('https://www.fisg.it/upload/result/6869/online',
    #        '2025_Italian_Champs', write_to_database=False, pdf_folder=pdf_folder, year=2425, isFSM=True)
    # scrape('https://www.isuresults.com/results/season2324/fc2024',
    #        '2024_Four_Continents', write_to_database=False, pdf_folder=pdf_folder, year=2425, isFSM=True)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2025/36369',
    #        '2025_Cranberry_Cup', write_to_database=False, pdf_folder=pdf_folder, year=2425, isFSM=True)
    # scrape('https://ijs.usfigureskating.org/leaderboard/results/2026/36273',
    #        '2026_US_Championships_Senior', write_to_database=True, pdf_folder=pdf_folder, year=2526, isFSM=True)
    scrape(
        "https://ijs.usfigureskating.org/leaderboard/results/2022/30702",
        "2025_Porter",
        write_to_database=True,
        pdf_folder=pdf_folder,
        excel_folder=excel_folder,
        year="2526",
        isFSM=False,
    )
    loadInfoForExistingCompetitions()
